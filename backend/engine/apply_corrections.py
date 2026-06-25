"""
Apply artist-name corrections from an Issues file to its matching
Annotated copy, then log every change.

Workflow:
    1. The user fills in the "Correction" column on the Artist Clusters tab
       of <name>_issues.xlsx. The column is pre-filled with the scanner's
       canonical guess, so most rows just need a glance.
       - Empty cell  → skip this cluster.
       - A name      → replace every variant in the cluster with this name
                       across all artist-bearing columns of the annotated
                       copy.
       - "LEAVE"     → mark the cluster as intentional. Future scans won't
                       flag it again (e.g., the DeBarge brothers).
    2. Run this script with <name>_annotated.xlsx as the argument
       (typically via "Apply Artist Corrections.command").
    3. The annotated copy is rewritten in place with the corrections applied.
    4. An "Applied Corrections" tab is appended/refreshed in the issues file
       with a row per cell that changed (sheet, row, column, before, after).
    5. LEAVE markers are saved to .artist_leave.json so the scanner stops
       flagging those clusters on subsequent scans.

Usage:
    python3 apply_corrections.py <annotated.xlsx>
    python3 apply_corrections.py <annotated.xlsx> --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Reuse the scanner's normalization, splitter, column discovery, and constants
# so we match exactly what the scanner clustered.
from scan_metadata import (
    DELIMITERS_RE,
    EXTRA_SINGLE_NAME_COLUMNS,
    LEAVE_MARKER,
    LEAVE_RECORDS_FILENAME,
    MAX_ARTIST_SLOTS,
    MULTI_NAME_COLUMNS,
    SINGLE_NAME_COLUMN_PATTERNS,
    detect_tracks_sheet,
    load_leave_records,
    normalize,
    parse_variants_str,
    save_leave_records,
)

# Same delimiter pattern as DELIMITERS_RE, but capturing — so re.split keeps
# the delimiters in the output and we can rejoin without losing original
# punctuation/spacing (e.g., "Foo, Bar feat. Baz" stays as written).
DELIMITERS_CAPTURE_RE = re.compile(
    r"(\s*[,/|;]\s*| feat\. | ft\. | & | x | vs )", flags=re.IGNORECASE
)


# ---- Reading the Correction column ----------------------------------------


def read_corrections(issues_path: Path) -> list[dict]:
    """
    Parse the Artist Clusters tab into a list of cluster records:
        {
            "cluster_id":  "T1",
            "canonical":   "Marcelo Rezende",
            "correction":  "Marcelo Rezende"  (or "" or "LEAVE"),
            "variants":    ["Marcelo Rezende", "Marceloa Rezende"]  (original case)
        }
    """
    wb = load_workbook(issues_path, read_only=True, data_only=True)
    if "Artist Clusters" not in wb.sheetnames:
        raise SystemExit(
            f"No 'Artist Clusters' tab in {issues_path.name}. "
            "Run a scan first so the issues file is generated."
        )
    ws = wb["Artist Clusters"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    needed = {"Cluster", "Canonical", "Correction", "Variants (count)"}
    missing = needed - set(header)
    if missing:
        raise SystemExit(
            f"'Artist Clusters' tab is missing column(s): {sorted(missing)}. "
            "Re-run the scan to regenerate the issues file with the new layout."
        )
    idx = {name: header.index(name) for name in needed}
    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        out.append({
            "cluster_id": str(row[idx["Cluster"]] or "").strip(),
            "canonical":  str(row[idx["Canonical"]] or "").strip(),
            "correction": str(row[idx["Correction"]] or "").strip(),
            "variants":   parse_variants_str(str(row[idx["Variants (count)"]] or "")),
        })
    return out


# ---- Building the replacement table ---------------------------------------


def build_replacement_table(
    clusters: list[dict],
) -> tuple[dict[str, str], list[dict], list[dict]]:
    """
    Returns:
        replacements   — { normalized_variant_name: target_name } for swap
        leave_records  — clusters the user marked with LEAVE (to persist)
        skipped        — clusters with empty Correction (no action)
    """
    replacements: dict[str, str] = {}
    leave_records: list[dict] = []
    skipped: list[dict] = []
    seen_now = datetime.now().strftime("%Y-%m-%d %H:%M")

    for c in clusters:
        corr = c["correction"]
        if not corr:
            skipped.append(c)
            continue
        if corr.upper() == LEAVE_MARKER:
            leave_records.append({
                "variants": c["variants"],
                "cluster_id": c["cluster_id"],
                "added_at": seen_now,
                "note": "Marked LEAVE via the Correction column.",
            })
            continue
        for variant in c["variants"]:
            key = normalize(variant)
            if not key:
                continue
            # If two clusters disagree on the same variant (very rare — the
            # scanner clusters by similarity), the later one wins. Both will
            # show up in the audit log so it's traceable.
            replacements[key] = corr

    return replacements, leave_records, skipped


# ---- Applying replacements to a sheet -------------------------------------


def replace_in_cell(value: str, replacements: dict[str, str]) -> tuple[str, list[tuple[str, str]]]:
    """
    Replace any matching artist names inside a cell value, preserving the
    original delimiters (commas, ampersands, "feat.", etc.). Returns the new
    value and a list of (old_segment, new_segment) pairs that were swapped.
    Single-name cells are handled by passing through this same function — a
    cell with no delimiters is just one segment.
    """
    if not isinstance(value, str) or not value:
        return value, []
    parts = DELIMITERS_CAPTURE_RE.split(value)
    out_parts: list[str] = []
    changes: list[tuple[str, str]] = []
    for i, p in enumerate(parts):
        if i % 2 == 1:  # delimiter chunk — leave it alone
            out_parts.append(p)
            continue
        if not p:
            out_parts.append(p)
            continue
        # Preserve leading/trailing whitespace from the original segment so
        # we don't accidentally "tighten" the cell.
        stripped = p.strip()
        if not stripped:
            out_parts.append(p)
            continue
        new_name = replacements.get(normalize(stripped))
        if new_name and new_name != stripped:
            leading = p[: len(p) - len(p.lstrip())]
            trailing = p[len(p.rstrip()):]
            out_parts.append(leading + new_name + trailing)
            changes.append((stripped, new_name))
        else:
            out_parts.append(p)
    return "".join(out_parts), changes


def apply_to_workbook(
    annotated_path: Path,
    replacements: dict[str, str],
) -> list[dict]:
    """
    Open the annotated workbook, find the tracks sheet, walk every artist-
    bearing column (multi- and single-name), apply replacements in place,
    and return an audit log.
    """
    if not replacements:
        return []

    tracks_sheet, _ = detect_tracks_sheet(annotated_path)
    wb = load_workbook(annotated_path)
    if tracks_sheet not in wb.sheetnames:
        raise SystemExit(f"Tracks sheet {tracks_sheet!r} not found in workbook.")
    ws = wb[tracks_sheet]

    # Build the column index from the first row (matches scan_metadata.py's
    # discover_artist_columns logic, but reads directly from the worksheet).
    header_row = [
        (c.value if c.value is not None else "") for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    name_to_col_idx: dict[str, int] = {}
    for j, h in enumerate(header_row, start=1):
        if isinstance(h, str):
            name_to_col_idx[h] = j

    multi_cols = [name_to_col_idx[c] for c in MULTI_NAME_COLUMNS if c in name_to_col_idx]
    single_cols: list[int] = []
    for n in range(1, MAX_ARTIST_SLOTS + 1):
        for pat in SINGLE_NAME_COLUMN_PATTERNS:
            col_name = pat.format(n=n)
            if col_name in name_to_col_idx:
                single_cols.append(name_to_col_idx[col_name])
    for col_name in EXTRA_SINGLE_NAME_COLUMNS:
        if col_name in name_to_col_idx:
            single_cols.append(name_to_col_idx[col_name])

    changes_log: list[dict] = []

    # Walk every data row. .max_row covers any row openpyxl knows about.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column not in multi_cols and cell.column not in single_cols:
                continue
            new_val, swaps = replace_in_cell(cell.value, replacements)
            if not swaps:
                continue
            cell.value = new_val
            for old, new in swaps:
                changes_log.append({
                    "Sheet": tracks_sheet,
                    "Excel Row": cell.row,
                    "Column": header_row[cell.column - 1] or get_column_letter(cell.column),
                    "Before": old,
                    "After": new,
                })

    if changes_log:
        wb.save(annotated_path)
    return changes_log


# ---- Audit-log tab in the issues file -------------------------------------


HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")


def write_audit_log(
    issues_path: Path,
    changes_log: list[dict],
    leave_records: list[dict],
    skipped_count: int,
    annotated_path: Path,
) -> None:
    """Append/replace the 'Applied Corrections' tab on the issues file."""
    wb = load_workbook(issues_path)
    name = "Applied Corrections"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([f"Applied {len(changes_log)} cell change(s) at {stamp}"])
    ws["A1"].font = Font(bold=True, size=12, name="Arial")
    ws.append([f"Annotated copy: {annotated_path.name}"])
    ws.append([f"Skipped clusters (empty Correction): {skipped_count}"])
    ws.append([f"LEAVE markers recorded: {len(leave_records)}"])
    ws.append([])

    headers = ["Sheet", "Excel Row", "Column", "Before", "After"]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    for i, ch in enumerate(changes_log, start=1):
        ws.append([ch[h] for h in headers])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = ALT_FILL

    # Column widths roughly tuned for readability.
    widths = {1: 22, 2: 10, 3: 30, 4: 35, 5: 35}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    if leave_records:
        ws.append([])
        ws.append(["LEAVE markers added this run:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, name="Arial")
        ws.append(["Cluster ID", "Variants"])
        hdr = ws.max_row
        for cell in ws[hdr]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for r in leave_records:
            ws.append([r.get("cluster_id", ""), "; ".join(r.get("variants", []))])

    wb.save(issues_path)


# ---- Main -----------------------------------------------------------------


def derive_issues_path(annotated_path: Path) -> Path:
    """Bossa - Montse's Copy_annotated.xlsx → ..._issues.xlsx (sibling)."""
    stem = annotated_path.stem
    if not stem.endswith("_annotated"):
        raise SystemExit(
            "Expected an annotated file (filename should end in '_annotated.xlsx'). "
            f"Got: {annotated_path.name}"
        )
    base = stem[: -len("_annotated")]
    return annotated_path.with_name(base + "_issues.xlsx")


def merge_leave_records(project_dir: Path, new_records: list[dict]) -> int:
    """
    Add new LEAVE entries to .artist_leave.json. De-dupes by normalized
    variant set. Returns the number of new records actually added.
    """
    existing_sets = set(load_leave_records(project_dir))  # list of frozensets
    # Reconstruct existing records as full dicts for re-saving.
    path = project_dir / LEAVE_RECORDS_FILENAME
    existing_records: list[dict] = []
    if path.exists():
        try:
            existing_records = json.loads(path.read_text()).get("records", [])
        except Exception:
            existing_records = []

    added = 0
    for rec in new_records:
        norm_set = frozenset(normalize(n) for n in rec.get("variants", []) if n)
        if not norm_set or norm_set in existing_sets:
            continue
        existing_records.append({
            "variants": rec.get("variants", []),
            "cluster_id": rec.get("cluster_id", ""),
            "note": rec.get("note", ""),
            "added_at": rec.get("added_at", ""),
        })
        existing_sets.add(norm_set)
        added += 1

    save_leave_records(project_dir, existing_records)
    return added


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Apply Correction column entries from <name>_issues.xlsx to <name>_annotated.xlsx."
    )
    parser.add_argument("annotated", help="Path to the <name>_annotated.xlsx file.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Compute changes but don't write any files. Useful for previewing.")
    args = parser.parse_args()

    annotated_path = Path(args.annotated).expanduser().resolve()
    if not annotated_path.exists():
        print(f"File not found: {annotated_path}", file=sys.stderr)
        return 1

    issues_path = derive_issues_path(annotated_path)
    if not issues_path.exists():
        print(
            f"Issues file not found: {issues_path}\n"
            "Run a scan first so the corrections worksheet is generated.",
            file=sys.stderr,
        )
        return 1

    clusters = read_corrections(issues_path)
    if not clusters:
        print("No artist clusters found in the issues file — nothing to do.")
        return 0

    replacements, leave_records, skipped = build_replacement_table(clusters)

    print(f"Clusters total: {len(clusters)}")
    print(f"  → with corrections to apply: {len({normalize(v): r for v, r in replacements.items()})}")
    print(f"  → marked LEAVE this run:     {len(leave_records)}")
    print(f"  → skipped (empty Correction):{len(skipped)}")

    if args.dry_run:
        print("\n[dry-run] would replace these (variant → target):")
        for k, v in sorted(replacements.items()):
            print(f"   {k!r:<40} → {v!r}")
        return 0

    changes_log = apply_to_workbook(annotated_path, replacements)
    print(f"Cell changes written to {annotated_path.name}: {len(changes_log)}")

    write_audit_log(issues_path, changes_log, leave_records,
                    skipped_count=len(skipped),
                    annotated_path=annotated_path)
    print(f"Audit log written to {issues_path.name} → 'Applied Corrections' tab.")

    if leave_records:
        # Store LEAVE markers at the project root (next to apply_corrections.py)
        # so a single decision applies to every phase folder.
        project_root = Path(__file__).resolve().parent
        added = merge_leave_records(project_root, leave_records)
        print(f"LEAVE markers stored in .artist_leave.json (newly added: {added}).")

    return 0


if __name__ == "__main__":
    sys.exit(main())
