"""
Apply format-validation fixes from an Issues file to its matching Annotated copy.

This handles all four format sub-checks in one pass:

  1. ISRC and UPC format errors → read the "Format Corrections" tab. Each
     row with a non-empty Corrected Value becomes a write to the
     (Excel Row, Column) cell on the annotated copy.

  2. Master Splits not summing to 100 → read the "Master Splits Review"
     tab. Each Artist N Master Split column is editable; whatever value
     the user has on each row gets written back to that exact cell.
     (No-op writes — i.e., values the user didn't change — still happen
     but are harmless.)

  3. Master Split columns formatted as % → for any column flagged as a
     column-wide percent-format problem, strip the percent number_format
     from the entire column so 0.25 displays as 0.25 (not 25%) and the
     ingestion picks up the raw value. Done automatically; no UI input.

  4. Audit log written to "Applied Format Fixes" tab in the issues file.

Usage:
    python3 apply_format_corrections.py <annotated.xlsx>
    python3 apply_format_corrections.py <annotated.xlsx> --dry-run
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scan_metadata import detect_tracks_sheet


# ---- Reading the review tabs ---------------------------------------------


def read_format_corrections(issues_path: Path) -> list[dict]:
    """
    Parse the Format Corrections tab → list of cell-correction records.
    Records with an empty Corrected Value are returned too (so the caller
    can count "deferred" rows); the apply step filters them out.
    """
    wb = load_workbook(issues_path, read_only=True, data_only=True)
    if "Format Corrections" not in wb.sheetnames:
        return []
    ws = wb["Format Corrections"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    needed = {"Type", "Excel Row", "Column", "Found Value", "Corrected Value"}
    missing = needed - set(header)
    if missing:
        raise SystemExit(
            f"'Format Corrections' tab is missing column(s): {sorted(missing)}. "
            "Re-run the scan to regenerate the issues file."
        )
    idx = {n: header.index(n) for n in needed}
    out: list[dict] = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        try:
            excel_row = int(r[idx["Excel Row"]])
        except (TypeError, ValueError):
            continue
        colname = str(r[idx["Column"]] or "").strip()
        if not colname:
            continue
        out.append({
            "type":        str(r[idx["Type"]] or "").strip(),
            "excel_row":   excel_row,
            "column":      colname,
            "found":       "" if r[idx["Found Value"]] is None else str(r[idx["Found Value"]]).strip(),
            "corrected":   "" if r[idx["Corrected Value"]] is None else str(r[idx["Corrected Value"]]).strip(),
        })
    return out


def read_splits_review(issues_path: Path) -> list[dict]:
    """
    Parse the Master Splits Review tab → list of split-row records:
        {"excel_row": 849, "splits": {"Artist 1 Master Split": 0, "Artist 2 Master Split": 75, ...}}
    """
    wb = load_workbook(issues_path, read_only=True, data_only=True)
    if "Master Splits Review" not in wb.sheetnames:
        return []
    ws = wb["Master Splits Review"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    if "Excel Row" not in header:
        return []
    i_row = header.index("Excel Row")
    split_col_idxs = [
        (j, h) for j, h in enumerate(header)
        if h.startswith("Artist ") and h.endswith(" Master Split")
    ]
    out: list[dict] = []
    for r in rows[1:]:
        if not r or i_row >= len(r):
            continue
        try:
            excel_row = int(r[i_row])
        except (TypeError, ValueError):
            continue
        splits: dict[str, object] = {}
        for j, name in split_col_idxs:
            if j >= len(r):
                continue
            v = r[j]
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            splits[name] = v
        if splits:
            out.append({"excel_row": excel_row, "splits": splits})
    return out


def read_pct_columns(issues_path: Path) -> list[str]:
    """
    Read the column-wide format-issue rows from the Format Issues tab and
    return the list of column names that were flagged as having % formatting.
    """
    wb = load_workbook(issues_path, read_only=True, data_only=True)
    if "Format Issues" not in wb.sheetnames:
        return []
    ws = wb["Format Issues"]
    rows = list(ws.iter_rows(values_only=True))
    pct_columns: list[str] = []
    for r in rows:
        if not r:
            continue
        # The column-wide section has rows like
        #   ("Master split column with % formatting", "Artist 1 Master Split", ...)
        # The first cell is the Issue label. We match on it loosely.
        first = str(r[0] or "")
        if "% formatting" in first.lower() or "with % format" in first.lower():
            col_name = str(r[1] or "").strip() if len(r) > 1 else ""
            if col_name:
                pct_columns.append(col_name)
    return pct_columns


# ---- Writing the fixes ---------------------------------------------------


def _coerce_split_value(v):
    """Turn a user-entered split value into a clean number (int when whole)."""
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return int(v)
        return v
    s = str(v).strip().replace("%", "")
    if not s:
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s  # let openpyxl write the raw string; user can clean up later


def apply_to_workbook(
    annotated_path: Path,
    cell_corrections: list[dict],
    split_rows: list[dict],
    pct_columns: list[str],
) -> tuple[list[dict], list[dict], list[str]]:
    """
    Apply all format fixes to the annotated workbook in one save.
    Returns (cell_audit, split_audit, columns_format_stripped).
    """
    if not cell_corrections and not split_rows and not pct_columns:
        return [], [], []

    tracks_sheet, _ = detect_tracks_sheet(annotated_path)
    wb = load_workbook(annotated_path)
    if tracks_sheet not in wb.sheetnames:
        raise SystemExit(f"Tracks sheet {tracks_sheet!r} not found in workbook.")
    ws = wb[tracks_sheet]

    # Header → 1-based column index
    header_row = [
        (c.value if c.value is not None else "")
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    col_to_idx: dict[str, int] = {}
    for j, h in enumerate(header_row, start=1):
        if isinstance(h, str):
            col_to_idx[h] = j

    # 1) ISRC / UPC format corrections (only those with non-empty Corrected Value)
    cell_audit: list[dict] = []
    for f in cell_corrections:
        if not f["corrected"]:
            continue
        col_idx = col_to_idx.get(f["column"])
        if col_idx is None:
            continue
        cell = ws.cell(row=f["excel_row"], column=col_idx)
        before = "" if cell.value is None else str(cell.value)
        cell.value = f["corrected"]
        cell_audit.append({
            "Type":      f["type"],
            "Sheet":     tracks_sheet,
            "Excel Row": f["excel_row"],
            "Column":    f["column"],
            "Before":    before,
            "After":     f["corrected"],
        })

    # 2) Master split row writes
    split_audit: list[dict] = []
    for sr in split_rows:
        for col_name, raw_val in sr["splits"].items():
            col_idx = col_to_idx.get(col_name)
            if col_idx is None:
                continue
            cell = ws.cell(row=sr["excel_row"], column=col_idx)
            before = cell.value
            new_val = _coerce_split_value(raw_val)
            cell.value = new_val
            # If the cell came in with % number_format, strip it so 25 displays as 25.
            nf = (cell.number_format or "")
            if "%" in nf:
                cell.number_format = "General"
            split_audit.append({
                "Sheet":     tracks_sheet,
                "Excel Row": sr["excel_row"],
                "Column":    col_name,
                "Before":    "" if before is None else str(before),
                "After":     "" if new_val is None else str(new_val),
            })

    # 3) Auto-strip % format from column-wide flagged columns. Walk every row
    # in those columns and replace any percent number_format with "General".
    columns_stripped: list[str] = []
    for col_name in pct_columns:
        col_idx = col_to_idx.get(col_name)
        if col_idx is None:
            continue
        touched = 0
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if "%" in (cell.number_format or ""):
                cell.number_format = "General"
                touched += 1
        if touched:
            columns_stripped.append(f"{col_name} ({touched} cells)")

    if cell_audit or split_audit or columns_stripped:
        wb.save(annotated_path)
    return cell_audit, split_audit, columns_stripped


# ---- Audit log -----------------------------------------------------------


HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")


def write_audit_log(
    issues_path: Path,
    cell_audit: list[dict],
    split_audit: list[dict],
    columns_stripped: list[str],
    annotated_path: Path,
) -> None:
    """Append/replace the 'Applied Format Fixes' tab on the issues file."""
    wb = load_workbook(issues_path)
    name = "Applied Format Fixes"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([f"Applied at {stamp}"])
    ws["A1"].font = Font(bold=True, size=12, name="Arial")
    ws.append([f"Annotated copy: {annotated_path.name}"])
    ws.append([f"ISRC/UPC corrections: {len(cell_audit)}   "
               f"Master split writes: {len(split_audit)}   "
               f"Columns %-stripped: {len(columns_stripped)}"])
    ws.append([])

    if cell_audit:
        ws.append(["ISRC / UPC corrections:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, name="Arial")
        headers = ["Type", "Sheet", "Excel Row", "Column", "Before", "After"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for i, c in enumerate(cell_audit, start=1):
            ws.append([c[h] for h in headers])
            if i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = ALT_FILL
        ws.append([])

    if split_audit:
        ws.append(["Master split writes:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, name="Arial")
        headers = ["Sheet", "Excel Row", "Column", "Before", "After"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for i, c in enumerate(split_audit, start=1):
            ws.append([c[h] for h in headers])
            if i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = ALT_FILL
        ws.append([])

    if columns_stripped:
        ws.append(["Columns where % format was stripped:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, name="Arial")
        for col in columns_stripped:
            ws.append([col])

    widths = {1: 16, 2: 22, 3: 10, 4: 24, 5: 22, 6: 22}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(issues_path)


# ---- Main ----------------------------------------------------------------


def derive_issues_path(annotated_path: Path) -> Path:
    """<name>_annotated.xlsx → <name>_issues.xlsx (sibling)."""
    stem = annotated_path.stem
    if not stem.endswith("_annotated"):
        raise SystemExit(
            "Expected an annotated file (filename should end in '_annotated.xlsx'). "
            f"Got: {annotated_path.name}"
        )
    base = stem[: -len("_annotated")]
    return annotated_path.with_name(base + "_issues.xlsx")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Apply Format Corrections + Master Splits Review entries from "
            "<name>_issues.xlsx to <name>_annotated.xlsx, and auto-strip "
            "percent number_format from any column flagged as column-wide %."
        )
    )
    parser.add_argument("annotated", help="Path to the <name>_annotated.xlsx file.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Compute changes but don't write any files.")
    args = parser.parse_args()

    annotated_path = Path(args.annotated).expanduser().resolve()
    if not annotated_path.exists():
        print(f"File not found: {annotated_path}", file=sys.stderr)
        return 1

    issues_path = derive_issues_path(annotated_path)
    if not issues_path.exists():
        print(
            f"Issues file not found: {issues_path}\n"
            "Run a scan first so the corrections worksheet is generated.",
            file=sys.stderr,
        )
        return 1

    cell_corrections = read_format_corrections(issues_path)
    split_rows = read_splits_review(issues_path)
    pct_columns = read_pct_columns(issues_path)

    fills_to_apply = sum(1 for f in cell_corrections if f["corrected"])
    deferred = sum(1 for f in cell_corrections if not f["corrected"])
    split_writes = sum(len(sr["splits"]) for sr in split_rows)

    print(f"Format Corrections rows total:  {len(cell_corrections)}")
    print(f"  → corrections to apply:       {fills_to_apply}")
    print(f"  → deferred (empty):           {deferred}")
    print(f"Master split row reviews:       {len(split_rows)} ({split_writes} cell write(s))")
    print(f"Columns to %-strip:             {len(pct_columns)}  {pct_columns}")

    if args.dry_run:
        print("\n[dry-run] would change these cells:")
        for f in cell_corrections:
            if f["corrected"]:
                print(f"   row {f['excel_row']:>4}  {f['column']:<28}  {f['found']!r} → {f['corrected']!r}")
        for sr in split_rows:
            print(f"   row {sr['excel_row']:>4}  splits: {sr['splits']}")
        if pct_columns:
            print(f"   strip % format from columns: {pct_columns}")
        return 0

    cell_audit, split_audit, columns_stripped = apply_to_workbook(
        annotated_path, cell_corrections, split_rows, pct_columns
    )
    print(f"\nWrites to {annotated_path.name}: "
          f"{len(cell_audit)} ISRC/UPC, {len(split_audit)} split, "
          f"{len(columns_stripped)} column(s) %-stripped")

    write_audit_log(issues_path, cell_audit, split_audit, columns_stripped, annotated_path)
    print(f"Audit log written to {issues_path.name} → 'Applied Format Fixes' tab.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
