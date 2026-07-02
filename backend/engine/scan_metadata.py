"""
Metadata Sheet Scanner — v3
============================
Scans a Label Engine metadata sheet for issues across three supported formats:
  • label-engine   — Label Engine Master Metadata template (full ingestion sheet)
  • internal-deals — Internal Deals template (master + composition tabs; trimmed fields)
  • epicwin-splits — EpicWin splits-correction import (ISRC/splits/ID checks only)

Legacy description:
Scans a Bossa-format metadata sheet for two kinds of issues:
    1. Misspelled / inconsistent artist names (fuzzy matching).
    2. Duplicate ISRCs that appear on rows with different artists.

Two output modes (or both):
    "report"    — separate Issues.xlsx with one row per problem (default).
    "annotated" — copy of the original sheet with problem cells highlighted and
                  Excel comments explaining each issue.
    "both"      — produce both files in one run.

Usage:
    python scan_metadata.py "<path-to-metadata.xlsx>"
    python scan_metadata.py "<path-to-metadata.xlsx>" --mode annotated
    python scan_metadata.py "<path-to-metadata.xlsx>" --mode both

Tunable constants (top of file):
    SIMILARITY_THRESHOLD — how close two names must be to be considered the same artist.
    MIN_NAME_LENGTH      — names shorter than this are ignored (e.g., "Cris" is a nickname, not a typo).
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

# ---- Tunable settings -------------------------------------------------------

SIMILARITY_THRESHOLD = 85   # 0–100. Higher = stricter (fewer flags, fewer false positives).
MIN_NAME_LENGTH = 4         # Skip very short names like "Cris" / "Nina" — too noisy to fuzzy-match.

# Track-sheet columns whose values are lists of artists (split on these delimiters).
DELIMITERS_RE = re.compile(r"\s*[,/|;]\s*| feat\. | ft\. | & | x | vs ", flags=re.IGNORECASE)

# Sheet auto-detection: the scanner picks the worksheet that has a "Track ISRC"
# column AND the most "core metadata" columns. This lets it work on workbooks that
# happen to have Track ISRC in multiple tabs (e.g., summary views, upload-formatted
# views, contract-tracking sheets, etc.) without picking the wrong one.
CORE_TRACK_COLUMNS = [
    "Track ISRC", "UPC", "Track Title", "Track Display Artist",
    "Release Artist", "Release Title", "Release Date", "Release Type",
    "Track P Line", "Track C Line", "Track Genre", "Track Language",
    "Track Explicit", "Audio File Path", "Artwork File Path",
    "Artist 1 Client Name", "Artist 1 Name on Track", "Artist 1 Master Split",
    "Artist 1 Contract Name", "Artist 1 Real Name(s) (comma separated)",
]
# Roster sheet (only present in some layouts; not currently consumed by checks).
ROSTER_SHEET_NAME = "Artist Check List"

# Columns to scan. Track Display Artist + Release Artist may contain multiple names; the
# Artist N columns each hold a single name. We accept multiple naming variants per slot
# because different label templates use slightly different headers (e.g., Bossa uses
# "Artist N Name on Track" while Mau5trap uses "Artist N Client Name On Track" for the
# first few slots).
MULTI_NAME_COLUMNS = ["Release Artist", "Track Display Artist"]
# Single-name artist columns scanned for typos. We deliberately exclude
# "Artist N Contract Name" — labels frequently use that column as a contract
# identifier ("<artist> - <album>" or "<artist> - <song>") rather than a
# clean artist name, which causes the fuzzy matcher to mega-cluster songs
# that share a common prefix/suffix.
SINGLE_NAME_COLUMN_PATTERNS = [
    "Artist {n} Client Name",
    "Artist {n} Name on Track",
    "Artist {n} Client Name On Track",   # Mau5trap variant for early slots
]
MAX_ARTIST_SLOTS = 12
# Other single-name columns to typo-check (no slot suffix). Label Name is
# included because typos in label names slip through easily and there are
# usually only a handful of distinct values per sheet, so cluster noise is
# minimal.
EXTRA_SINGLE_NAME_COLUMNS = ["Label Name"]

# Columns used by the duplicate-ISRC check.
ISRC_COLUMN = "Track ISRC"
TRACK_TITLE_COLUMN = "Track Title"
DISPLAY_ARTIST_COLUMN = "Track Display Artist"

# Issue types — used in the "Type" column of the Issues sheet.
TYPE_ARTIST = "Artist typo"
TYPE_ISRC = "Duplicate ISRC"
TYPE_MISSING = "Missing required field"
TYPE_FORMAT = "Format issue"

# Format-check settings.
UPC_COLUMN = "UPC"
SPLIT_PCT_COLUMN_THRESHOLD = 0.8   # % of non-empty cells with %-format → treat as column-wide.
SPLIT_SUM_TARGET = 100.0
SPLIT_SUM_TOLERANCE = 0.5          # accept 99.5–100.5 due to rounding.

# ISRC: 2 letters (country) + 3 alphanumeric (registrant) + 7 digits (year + designation).
# Hyphens stripped before matching.
ISRC_PATTERN = re.compile(r"^[A-Z]{2}[A-Z0-9]{3}[0-9]{7}$")
# UPC: 12 (UPC-A) or 13 (EAN-13) digits.
UPC_PATTERN = re.compile(r"^[0-9]{12,13}$")

# Required fields. Each row must have a non-empty, non-placeholder value in each entry.
# - A string entry means "this exact column is required."
# - A tuple entry means "any of these columns is acceptable" — used when different
#   label templates name the same field differently.
# (Genre is intentionally not in this list — see project history.)
REQUIRED_FIELDS: list[str | tuple[str, ...]] = [
    "UPC",
    "Release Title",
    "Release Artist",
    "Release Date",
    "Release Type",
    "Track ISRC",
    "Track Title",
    "Track Display Artist",
    "Track Number",
    "Track P Line",
    "Track C Line",
    "Track Language",
    "Track Explicit",
    "Audio File Path",
    "Artwork File Path",
    "Artist 1 Client Name",
    # Bossa uses "Artist 1 Name on Track" (lowercase t).
    # Mau5trap uses "Artist 1 Client Name On Track" (capital T) for the first few slots.
    ("Artist 1 Name on Track", "Artist 1 Client Name On Track"),
]

# Strings (case-insensitive, after stripping) that count as a placeholder rather than
# real data. Treated the same as a blank for the missing-required-fields check.
PLACEHOLDER_VALUES = {
    "n/a", "n.a.", "na",
    "tbd", "t.b.d.",
    "tba", "t.b.a.",
    "unknown", "none", "null", "nil",
    "?", "-", "--", "---",
}


# ---- Format schemas ---------------------------------------------------------
#
# Each schema declares:
#   display_name   — shown in the UI and report headers
#   sheet_hints    — ordered list of sheet-name substrings to look for (first match wins)
#   column_aliases — maps the sheet's column names → canonical engine column names
#                    Applied as a rename step before any checks run, so all existing
#                    logic stays untouched.
#   required_fields_override — if not None, replaces REQUIRED_FIELDS for this format
#   checks         — set of check names to run; omit a name to skip that check
#                    Recognised names: "artist_typos", "isrc_duplicates",
#                    "missing_fields", "format_validation",
#                    "splits_correction"  (EpicWin-specific)
#
# The "label-engine" schema is the identity mapping — no aliases, all checks,
# REQUIRED_FIELDS unchanged — kept explicit so detect_format() has a concrete
# object to return.

FORMAT_SCHEMAS: dict[str, dict] = {
    "label-engine": {
        "display_name": "Label Engine Master",
        "sheet_hints": ["Metadata - Master"],
        "column_aliases": {},
        "required_fields_override": None,   # use global REQUIRED_FIELDS
        "checks": {"artist_typos", "isrc_duplicates", "missing_fields", "format_validation"},
    },
    "internal-deals-master": {
        "display_name": "Internal Deals — Master",
        "sheet_hints": ["Metadata - Master (Internal", "Metadata - Master"],
        "column_aliases": {
            # The internal deals sheet uses "ISRC" and "Label" instead of the
            # canonical engine names.
            "ISRC": "Track ISRC",
            "Label": "Label Name",
        },
        # This is a summary / deal sheet — audio/artwork paths and full artist
        # slots are intentionally absent. Only check what's actually present.
        "required_fields_override": [
            "UPC",
            "Release Title",
            "Track ISRC",
            "Track Title",
            "Track Display Artist",
            "Release Date",
        ],
        "checks": {"artist_typos", "isrc_duplicates", "missing_fields", "format_validation"},
    },
    "internal-deals-composition": {
        "display_name": "Internal Deals — Composition",
        "sheet_hints": ["Metadata - Composition (Interna", "Metadata - Composition"],
        "column_aliases": {
            "ISRC": "Track ISRC",
            "Label": "Label Name",
        },
        "required_fields_override": [
            "UPC",
            "Release Title",
            "Track ISRC",
            "Track Title",
            "Track Display Artist",
            "Release Date",
        ],
        "checks": {"artist_typos", "isrc_duplicates", "missing_fields", "format_validation"},
    },
    "epicwin-splits": {
        "display_name": "Splits Correction",
        "sheet_hints": ["Sheet1", "Splits"],
        "column_aliases": {
            # The "NEW " prefix is stripped before aliasing (see normalize_columns).
            # These are the canonical post-strip → engine-canonical mappings.
            "ISRC": "Track ISRC",
            "Track Artist": "Track Display Artist",
            "Client or Label ID": "_splits_client_id",
            "Client Name": "_splits_client_name",
            "Account ID": "_splits_account_id",
            "Account Name": "_splits_account_name",
            "Allocation Percentage": "_splits_allocation_pct",
            "Allocation Type": "_splits_allocation_type",
            "Net / Gross": "_splits_net_gross",
        },
        "required_fields_override": [],   # no required-fields check for splits sheets
        "checks": {"artist_typos", "isrc_duplicates", "format_validation", "splits_correction"},
    },
}

# Fingerprint columns used to identify each format from the header row.
# A format matches if ≥ MATCH_THRESHOLD of its fingerprint columns are present.
FORMAT_FINGERPRINTS: dict[str, list[str]] = {
    "label-engine": [
        "Label Name", "Track ISRC", "UPC", "Release Artist", "Release Title",
        "Artist 1 Client Name", "Artist 1 Name on Track",
    ],
    "internal-deals-master": [
        "ISRC", "UPC", "Release Title", "Track Title", "Track Display Artist",
        "Current LE Royalty Split - Seller",
    ],
    "internal-deals-composition": [
        "ISRC", "UPC", "Release Title", "Track Title", "ISWC",
        "Current LE Royalty Split - Seller",
    ],
    "epicwin-splits": [
        "ISRC", "UPC", "Release Artist", "Track Artist", "Track Title",
        "Allocation Percentage", "Net / Gross",
    ],
}
FORMAT_MATCH_THRESHOLD = 0.6   # fraction of fingerprint columns that must be present


def _strip_new_prefix(headers: list[str | None]) -> list[str | None]:
    """Strip the 'NEW ' prefix from EpicWin-style column headers."""
    out = []
    for h in headers:
        if isinstance(h, str) and h.upper().startswith("NEW "):
            out.append(h[4:].strip())
        else:
            out.append(h)
    return out


def detect_format(wb, sheet_name: str) -> dict:
    """
    Detect which FORMAT_SCHEMAS entry best matches the given sheet.
    Returns the schema dict (never None — falls back to "label-engine").
    """
    ws = wb[sheet_name]
    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
    headers = set(str(h).strip() for h in _strip_new_prefix(raw_headers) if h is not None)

    best_key = "label-engine"
    best_score = 0.0
    for key, fingerprint in FORMAT_FINGERPRINTS.items():
        if not fingerprint:
            continue
        score = sum(1 for col in fingerprint if col in headers) / len(fingerprint)
        if score > best_score:
            best_score = score
            best_key = key

    if best_score < FORMAT_MATCH_THRESHOLD:
        best_key = "label-engine"

    return {**FORMAT_SCHEMAS[best_key], "_key": best_key}


def normalize_columns(df: pd.DataFrame, schema: dict) -> pd.DataFrame:
    """
    Rename columns to canonical engine names according to the schema's
    column_aliases map. Also strips the 'NEW ' prefix from any remaining
    headers (EpicWin files may have inconsistent prefixing).
    Returns a new DataFrame with renamed columns; the original is unchanged.
    """
    # Step 1: strip "NEW " prefix from all column names.
    new_cols = {}
    for col in df.columns:
        if isinstance(col, str) and col.upper().startswith("NEW "):
            new_cols[col] = col[4:].strip()
    if new_cols:
        df = df.rename(columns=new_cols)

    # Step 2: apply schema aliases.
    aliases = schema.get("column_aliases", {})
    if aliases:
        df = df.rename(columns={k: v for k, v in aliases.items() if k in df.columns})

    return df


# ---- EpicWin splits-specific checks -----------------------------------------

TYPE_SPLITS = "Splits error"
TYPE_ID_MISMATCH = "ID mismatch"


def find_splits_correction_issues(
    df: pd.DataFrame,
    sheet_name: str,
) -> tuple[list[dict], list[dict], list[dict]]:
    """
    EpicWin splits-correction checks:
      1. Split sum per ISRC must equal 100 (within SPLIT_SUM_TOLERANCE).
      2. Client or Label ID must map 1-to-1 with Client Name.
      3. Account ID must map 1-to-1 with Account Name.

    Returns (issues, split_errors, id_mismatches).
      issues        — unified list for the Issues tab
      split_errors  — per-ISRC rows for the Split Errors tab
      id_mismatches — per-conflict rows for the ID Mismatches tab
    """
    issues: list[dict] = []
    split_errors: list[dict] = []
    id_mismatches: list[dict] = []

    isrc_col = "Track ISRC" if "Track ISRC" in df.columns else None
    pct_col = "_splits_allocation_pct" if "_splits_allocation_pct" in df.columns else None
    track_id_col = "Track ID" if "Track ID" in df.columns else None
    title_col = "Track Title" if "Track Title" in df.columns else None
    client_id_col = "_splits_client_id" if "_splits_client_id" in df.columns else None
    client_name_col = "_splits_client_name" if "_splits_client_name" in df.columns else None
    account_id_col = "_splits_account_id" if "_splits_account_id" in df.columns else None
    account_name_col = "_splits_account_name" if "_splits_account_name" in df.columns else None

    # ---- 1. Splits sum to 100 (Net rows only) --------------------------------
    # Gross allocations (Net/Gross = "G") are variable and do not need to sum
    # to 100. Only Net rows ("N") must sum to 100.
    net_gross_col = "_splits_net_gross" if "_splits_net_gross" in df.columns else None

    release_id_col = "Release ID" if "Release ID" in df.columns else None

    if isrc_col and pct_col:
        # Group by Release ID + ISRC so the same track appearing on multiple
        # albums is checked independently per release (not summed together).
        grouped: dict[str, list[tuple[int, float]]] = defaultdict(list)
        for idx, row in df.iterrows():
            isrc_val = str(row.get(isrc_col, "") or "").strip()
            if not isrc_val:
                continue
            release_val = str(row.get(release_id_col, "") or "").strip() if release_id_col else ""
            key_val = f"{release_val}|{isrc_val}" if release_val else isrc_val
            # Skip Gross rows — only Net rows must sum to 100.
            if net_gross_col:
                ng = str(row.get(net_gross_col, "") or "").strip().upper()
                if ng == "G":
                    continue
            try:
                pct = float(row.get(pct_col, 0) or 0)
            except (ValueError, TypeError):
                pct = 0.0
            grouped[key_val].append((int(idx) + 2, pct))

        for group_key, row_pairs in sorted(grouped.items()):
            total = sum(p for _, p in row_pairs)
            diff = abs(total - SPLIT_SUM_TARGET)
            if diff <= SPLIT_SUM_TOLERANCE:
                continue
            # group_key is "releaseID|ISRC" or just "ISRC" when no release col.
            if "|" in group_key:
                display_release, display_isrc = group_key.split("|", 1)
            else:
                display_release, display_isrc = "", group_key
            rows_str = ", ".join(str(r) for r, _ in row_pairs)
            # Look up track title from the first matching row.
            title_val = ""
            if title_col:
                first_row_idx = row_pairs[0][0] - 2  # convert back to df index
                if first_row_idx in df.index:
                    title_val = str(df.at[first_row_idx, title_col] or "").strip()
            for excel_row, pct in row_pairs:
                issues.append({
                    "Type": TYPE_SPLITS,
                    "Sheet": sheet_name,
                    "Excel Row": excel_row,
                    "Column": pct_col.replace("_splits_", "").replace("_", " ").title(),
                    "Found Value": str(pct),
                    "Suggested Value": "Splits must sum to 100",
                    "Similarity": "",
                    "Cluster": f"S-{group_key}",
                    "Notes": (
                        f"ISRC {display_isrc}"
                        + (f" / Release {display_release}" if display_release else "")
                        + f": net splits total {total:.2f}% (expected 100). Rows: {rows_str}."
                    ),
                })
            split_errors.append({
                "Release ID": display_release,
                "ISRC": display_isrc,
                "Track Title": title_val,
                "Rows": rows_str,
                "Split Total": round(total, 4),
                "Difference from 100": round(total - 100, 4),
            })

    # ---- 2. Client ID ↔ Client Name consistency -----------------------------
    def _check_id_name_pair(id_col: str | None, name_col: str | None, label: str) -> None:
        if not id_col or not name_col:
            return
        id_to_names: dict[str, set[str]] = defaultdict(set)
        name_to_ids: dict[str, set[str]] = defaultdict(set)
        id_to_rows: dict[str, list[int]] = defaultdict(list)
        for idx, row in df.iterrows():
            id_val = str(row.get(id_col, "") or "").strip()
            name_val = str(row.get(name_col, "") or "").strip()
            # Skip blank/NaN IDs — a missing ID means no account exists yet,
            # not a mismatch. "nan" is pandas' NaN rendered as a string.
            if not id_val or id_val.lower() == "nan" or not name_val or name_val.lower() == "nan":
                continue
            id_to_names[id_val].add(name_val)
            name_to_ids[name_val].add(id_val)
            id_to_rows[id_val].append(int(idx) + 2)

        mismatch_id = 1
        for id_val, names in sorted(id_to_names.items()):
            if len(names) <= 1:
                continue
            rows_str = ", ".join(str(r) for r in id_to_rows[id_val])
            names_str = "; ".join(sorted(names))
            for excel_row in id_to_rows[id_val]:
                issues.append({
                    "Type": TYPE_ID_MISMATCH,
                    "Sheet": sheet_name,
                    "Excel Row": excel_row,
                    "Column": id_col.replace("_splits_", "").replace("_", " ").title(),
                    "Found Value": id_val,
                    "Suggested Value": "Resolve to a single name",
                    "Similarity": "",
                    "Cluster": f"ID-{mismatch_id}",
                    "Notes": f"{label} ID '{id_val}' maps to multiple names: {names_str}. Rows: {rows_str}.",
                })
            id_mismatches.append({
                "Type": f"{label} ID → multiple names",
                "ID": id_val,
                "Names found": names_str,
                "Rows": rows_str,
            })
            mismatch_id += 1

        for name_val, ids in sorted(name_to_ids.items()):
            if len(ids) <= 1:
                continue
            ids_str = "; ".join(sorted(ids))
            id_mismatches.append({
                "Type": f"{label} name → multiple IDs",
                "ID": ids_str,
                "Names found": name_val,
                "Rows": "—",
            })

    _check_id_name_pair(client_id_col, client_name_col, "Client")
    _check_id_name_pair(account_id_col, account_name_col, "Account")

    return issues, split_errors, id_mismatches


# ---- Helpers ----------------------------------------------------------------


def detect_tracks_sheet(input_path: Path) -> tuple[str, list[tuple[str, int, int]]]:
    """
    Identify which worksheet holds the track data.

    For Label Engine Master files: picks the sheet with "Track ISRC" and the
    most CORE_TRACK_COLUMNS (existing behaviour).

    For Internal Deals files: looks for the "Metadata - Master" tab first.
    For EpicWin files: looks for a sheet with ISRC + Allocation Percentage
    columns (after stripping any "NEW " prefix).

    Returns (chosen_sheet_name, candidates) — candidates is the same scored
    list as before (empty for non-LE-Master formats).
    """
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        # ---- Try format-schema sheet hints first ----------------------------
        for schema_key, schema in FORMAT_SCHEMAS.items():
            if schema_key in ("label-engine",):
                continue   # handled by the original scoring logic below
            for hint in schema.get("sheet_hints", []):
                for sheet_name in wb.sheetnames:
                    if hint.lower() in sheet_name.lower():
                        ws = wb[sheet_name]
                        raw = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
                        cleaned = set(str(h).strip() for h in _strip_new_prefix(raw) if h)
                        fp = FORMAT_FINGERPRINTS.get(schema_key, [])
                        if fp:
                            score = sum(1 for c in fp if c in cleaned) / len(fp)
                            if score >= FORMAT_MATCH_THRESHOLD:
                                return sheet_name, []

        # ---- Original LE-Master scoring logic --------------------------------
        candidates: list[tuple[str, int, int]] = []
        for name in wb.sheetnames:
            ws = wb[name]
            first_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
            # Accept both "Track ISRC" (LE Master) and bare "ISRC" (Internal Deals).
            header_set = set(str(h).strip() for h in first_row if h)
            if "Track ISRC" not in header_set and "ISRC" not in header_set:
                continue
            score = sum(1 for c in CORE_TRACK_COLUMNS if c in header_set)
            n_cols = sum(1 for h in first_row if h is not None)
            candidates.append((name, score, n_cols))

        if not candidates:
            return wb.sheetnames[0], []

        candidates.sort(key=lambda t: (-t[1], -t[2], t[0]))
        return candidates[0][0], candidates
    finally:
        wb.close()


def detect_all_sheets_to_scan(input_path: Path) -> list[tuple[str, dict]]:
    """
    Return a list of (sheet_name, schema) pairs covering every sheet that
    should be scanned in this workbook.

    For most files this is one entry.  For Internal Deals files both the
    master-recording and composition metadata tabs are returned so the caller
    can run checks on each and merge the results.

    Deduplication rules:
    - For label-engine format: only the single highest-scoring sheet is kept
      (mirrors the original detect_tracks_sheet behaviour).
    - For Internal Deals: master + composition are both kept if present.
    - For EpicWin: the one matching sheet is kept.
    - composition is preferred over master when the ISWC column is present.
    """
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        candidates: list[tuple[str, str, float]] = []   # (sheet, schema_key, score)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
            cleaned_headers = set(str(h).strip() for h in _strip_new_prefix(raw) if h)

            # Check composition before master — ISWC presence disambiguates.
            # We iterate in a fixed priority order rather than dict order.
            priority_order = [
                "internal-deals-composition",
                "internal-deals-master",
                "epicwin-splits",
                "label-engine",
            ]

            best_key = None
            best_score = 0.0
            for key in priority_order:
                fp = FORMAT_FINGERPRINTS.get(key, [])
                if not fp:
                    continue
                score = sum(1 for c in fp if c in cleaned_headers) / len(fp)
                # Break ties in favour of composition when ISWC is present.
                if key == "internal-deals-composition" and "ISWC" not in cleaned_headers:
                    continue
                if score >= FORMAT_MATCH_THRESHOLD and score > best_score:
                    best_score = score
                    best_key = key

            if best_key:
                candidates.append((sheet_name, best_key, best_score))

        if not candidates:
            first = wb.sheetnames[0]
            return [(first, {"_key": "unknown", "display_name": "unknown", "checks": set()})]

        # For label-engine: keep only the single best sheet.
        le_candidates = [(s, k, sc) for s, k, sc in candidates if k == "label-engine"]
        other_candidates = [(s, k, sc) for s, k, sc in candidates if k != "label-engine"]

        results: list[tuple[str, dict]] = []
        if le_candidates and not other_candidates:
            # Pure label-engine file — pick the highest-scoring sheet only.
            le_candidates.sort(key=lambda t: -t[2])
            s, k, _ = le_candidates[0]
            results = [(s, {**FORMAT_SCHEMAS[k], "_key": k})]
        else:
            # Internal Deals / EpicWin — include all identified sheets.
            for s, k, _ in other_candidates:
                results.append((s, {**FORMAT_SCHEMAS[k], "_key": k}))
            # Also include the best label-engine sheet if any survive alongside others
            # (shouldn't happen for these file types, but be safe).
            if le_candidates:
                le_candidates.sort(key=lambda t: -t[2])
                s, k, _ = le_candidates[0]
                results.append((s, {**FORMAT_SCHEMAS[k], "_key": k}))

        return results
    finally:
        wb.close()


def required_field_present(req, df_columns) -> str | None:
    """Given a string or tuple-of-strings requirement, return the first
    candidate that's present in df_columns, or None if none are."""
    candidates = (req,) if isinstance(req, str) else tuple(req)
    for c in candidates:
        if c in df_columns:
            return c
    return None


def strip_accents(s: str) -> str:
    """'Bárbara' -> 'Barbara' for matching purposes only."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def normalize(name: str) -> str:
    """Lowercase + accent-stripped + whitespace-collapsed key used for matching."""
    if not isinstance(name, str):
        return ""
    return re.sub(r"\s+", " ", strip_accents(name)).strip().lower()


def split_multi(value: str) -> list[str]:
    """Split a 'Track Display Artist' style cell into individual names."""
    if not isinstance(value, str):
        return []
    parts = DELIMITERS_RE.split(value)
    return [p.strip() for p in parts if p and p.strip()]


# ---- Carry-forward helpers for the artist-correction workflow -------------

LEAVE_MARKER = "LEAVE"           # User types this in the Correction column
LEAVE_RECORDS_FILENAME = ".artist_leave.json"  # Lives next to the input sheet

# ISRC-conflict review:
#   - The "Confirm OK?" column on the ISRC Conflicts tab accepts any of the
#     markers in CONFIRM_OK_MARKERS (case-insensitive). Typing one means "this
#     duplicate is intentional — don't flag it again."
#   - Confirmed conflicts persist in .isrc_leave.json at the project root,
#     mirroring the artist .artist_leave.json system.
CONFIRM_OK_MARKERS = {"ok", "yes", "y", "x", "✓", "true", "1"}
ISRC_LEAVE_RECORDS_FILENAME = ".isrc_leave.json"


def is_confirm_ok(value) -> bool:
    """Return True if a 'Confirm OK?' cell value counts as a confirmation."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in CONFIRM_OK_MARKERS


def parse_variants_str(s: str) -> list[str]:
    """Parse 'Foo (50); Bar (1)' back into ['Foo', 'Bar']."""
    if not isinstance(s, str) or not s.strip():
        return []
    out: list[str] = []
    for chunk in s.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        # Drop a trailing "(<digits>)" if present.
        m = re.match(r"^(.*?)\s*\(\d+\)\s*$", chunk)
        out.append(m.group(1).strip() if m else chunk)
    return [n for n in out if n]


def _norm_set(names) -> frozenset:
    """Normalized set of names (for comparing clusters across scans)."""
    return frozenset(normalize(n) for n in names if n)


def load_prior_corrections(issues_path: Path | None) -> list[dict]:
    """
    Read the Correction column from a previous _issues.xlsx so a re-scan
    doesn't lose the user's in-progress work. Returns a list of records like
    {"variants": frozenset of normalized names, "correction": str}.
    """
    if not issues_path or not issues_path.exists():
        return []
    try:
        wb = load_workbook(issues_path, read_only=True, data_only=True)
        if "Artist Clusters" not in wb.sheetnames:
            return []
        ws = wb["Artist Clusters"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h or "").strip() for h in rows[0]]
        try:
            i_var = header.index("Variants (count)")
        except ValueError:
            return []
        # If the prior file is from before this column existed, there's no
        # carry-forward data — return empty so the new file pre-fills with
        # canonicals for every cluster.
        if "Correction" not in header:
            return []
        i_corr = header.index("Correction")
        out: list[dict] = []
        for row in rows[1:]:
            if not row or i_var >= len(row):
                continue
            variants_str = row[i_var] or ""
            variants = parse_variants_str(str(variants_str))
            if not variants:
                continue
            corr = (row[i_corr] if i_corr < len(row) else "") or ""
            out.append({"variants": _norm_set(variants), "correction": str(corr).strip()})
        return out
    except Exception:
        # If the previous file is malformed, just start fresh — this is a
        # convenience feature, not a critical path.
        return []


def load_prior_isrc_actions(issues_path: Path | None) -> dict[tuple[str, int], dict]:
    """
    Read the ISRC Conflicts tab from a previous _issues.xlsx so a re-scan
    doesn't lose the user's in-progress 'Confirm OK?' / 'Corrected ISRC'
    entries. Returns a map keyed by (ISRC, Excel Row).
    Tolerates the legacy per-conflict layout (older issues files); in that
    case it just returns an empty dict.
    """
    if not issues_path or not issues_path.exists():
        return {}
    try:
        wb = load_workbook(issues_path, read_only=True, data_only=True)
        if "ISRC Conflicts" not in wb.sheetnames:
            return {}
        ws = wb["ISRC Conflicts"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        header = [str(h or "").strip() for h in rows[0]]
        # Legacy layout had no Excel Row column; nothing to carry forward.
        if "Excel Row" not in header or "ISRC" not in header:
            return {}
        i_isrc = header.index("ISRC")
        i_row = header.index("Excel Row")
        i_ok = header.index("Confirm OK?") if "Confirm OK?" in header else -1
        i_corr = header.index("Corrected ISRC") if "Corrected ISRC" in header else -1
        out: dict[tuple[str, int], dict] = {}
        for r in rows[1:]:
            if not r or i_isrc >= len(r) or i_row >= len(r):
                continue
            isrc = str(r[i_isrc] or "").strip()
            try:
                row_num = int(r[i_row])
            except (TypeError, ValueError):
                continue
            if not isrc:
                continue
            ok_val = r[i_ok] if 0 <= i_ok < len(r) else ""
            corr_val = r[i_corr] if 0 <= i_corr < len(r) else ""
            out[(isrc, row_num)] = {
                "confirm_ok": "" if ok_val is None else str(ok_val).strip(),
                "corrected_isrc": "" if corr_val is None else str(corr_val).strip(),
            }
        return out
    except Exception:
        return {}


def _find_prior_correction(cluster: list[str], prior: list[dict]) -> str | None:
    """
    Match a new cluster to a prior cluster by variant overlap. Returns the
    user's previously-typed Correction (which may be "" if they cleared it).
    Returns None if no overlap was found, signaling "use the canonical".
    """
    if not prior:
        return None
    new_set = _norm_set(cluster)
    if not new_set:
        return None
    best_overlap = 0
    best_corr: str | None = None
    for rec in prior:
        overlap = len(new_set & rec["variants"])
        if overlap > best_overlap:
            best_overlap = overlap
            best_corr = rec["correction"]
    return best_corr


def load_leave_records(project_dir: Path) -> list[frozenset]:
    """
    Read .artist_leave.json — clusters the user has marked as intentional
    (e.g., DeBarge brothers). Each record is a normalized variant set.
    """
    path = project_dir / LEAVE_RECORDS_FILENAME
    if not path.exists():
        return []
    try:
        import json as _json
        data = _json.loads(path.read_text())
        return [frozenset(rec.get("variants_normalized", [])) for rec in data.get("records", [])]
    except Exception:
        return []


def save_leave_records(project_dir: Path, records: list[dict]) -> None:
    """
    Write .artist_leave.json. `records` is a list of dicts with at least
    'variants' (the original-case variant list, preserved for human reading)
    and optionally 'note', 'added_at'. The normalized form is computed here.
    """
    import json as _json
    path = project_dir / LEAVE_RECORDS_FILENAME
    payload = {"records": []}
    for rec in records:
        variants = rec.get("variants") or []
        payload["records"].append({
            "variants": list(variants),
            "variants_normalized": sorted({normalize(n) for n in variants if n}),
            "note": rec.get("note", ""),
            "added_at": rec.get("added_at", ""),
        })
    path.write_text(_json.dumps(payload, indent=2, ensure_ascii=False))


def _matches_leave(cluster: list[str], leave_records: list[frozenset]) -> bool:
    """A cluster matches a LEAVE record if its variant set is a subset of one."""
    new_set = _norm_set(cluster)
    return any(new_set and new_set <= rec for rec in leave_records)


# ---- ISRC-conflict LEAVE system -------------------------------------------
#
# Same idea as the artist-cluster LEAVE system, but the signature of an ISRC
# "leave" record is a frozenset of (isrc, normalized_artist) pairs. That way:
#   - The same ISRC + same artist set = recognized as the same intentional
#     conflict and stays suppressed across rescans.
#   - If a new artist gets added to that ISRC later, it's a NEW conflict
#     constellation and the scanner re-flags it (intentional safety net).


def _isrc_signature(occs: list[dict]) -> frozenset:
    """Build a stable signature for an ISRC conflict from its occurrences."""
    return frozenset(
        (str(o.get("isrc", "")).strip().upper(), normalize(o.get("artist", "")))
        for o in occs
        if o.get("isrc")
    )


def load_isrc_leave_records(project_dir: Path) -> list[frozenset]:
    """Read .isrc_leave.json — conflicts confirmed as intentional duplicates."""
    path = project_dir / ISRC_LEAVE_RECORDS_FILENAME
    if not path.exists():
        return []
    try:
        import json as _json
        data = _json.loads(path.read_text())
        return [
            frozenset(tuple(pair) for pair in rec.get("signature", []))
            for rec in data.get("records", [])
        ]
    except Exception:
        return []


def save_isrc_leave_records(project_dir: Path, records: list[dict]) -> None:
    """
    Write .isrc_leave.json. Each input record should carry:
        - 'signature' (list of [isrc, normalized_artist] pairs)
        - 'isrc'      (the conflicting ISRC, for human reading)
        - 'rows'      (list of Excel row numbers — informational only)
        - 'note', 'added_at' (optional)
    """
    import json as _json
    path = project_dir / ISRC_LEAVE_RECORDS_FILENAME
    payload = {"records": []}
    for rec in records:
        sig = rec.get("signature") or []
        payload["records"].append({
            "isrc": rec.get("isrc", ""),
            "signature": [list(pair) for pair in sig],
            "rows": list(rec.get("rows", [])),
            "note": rec.get("note", ""),
            "added_at": rec.get("added_at", ""),
        })
    path.write_text(_json.dumps(payload, indent=2, ensure_ascii=False))


def _isrc_conflict_is_left(occs: list[dict], leave_records: list[frozenset]) -> bool:
    """A conflict is suppressed if its signature is a subset of any leave record."""
    sig = _isrc_signature(occs)
    return any(sig and sig <= rec for rec in leave_records)


def discover_artist_columns(df: pd.DataFrame) -> tuple[list[str], list[str]]:
    """Return (multi_name_columns, single_name_columns) actually present in the sheet."""
    multi = [c for c in MULTI_NAME_COLUMNS if c in df.columns]
    singles: list[str] = []
    for n in range(1, MAX_ARTIST_SLOTS + 1):
        for pat in SINGLE_NAME_COLUMN_PATTERNS:
            col = pat.format(n=n)
            if col in df.columns:
                singles.append(col)
    for col in EXTRA_SINGLE_NAME_COLUMNS:
        if col in df.columns:
            singles.append(col)
    return multi, singles


# ---- Cluster building -------------------------------------------------------


class NameOccurrence:
    __slots__ = ("name", "row", "column", "kind")

    def __init__(self, name: str, row: int, column: str, kind: str):
        self.name = name           # exact value found in the cell (or split-out piece)
        self.row = row             # 1-based Excel row number (already includes the header row)
        self.column = column       # column header name
        self.kind = kind           # "single" or "multi" — how it was extracted


def collect_occurrences(df: pd.DataFrame) -> list[NameOccurrence]:
    multi_cols, single_cols = discover_artist_columns(df)
    occurrences: list[NameOccurrence] = []

    for idx, row in df.iterrows():
        excel_row = int(idx) + 2  # +1 for header, +1 for 1-based indexing

        for col in multi_cols:
            val = row[col]
            for piece in split_multi(val):
                occurrences.append(NameOccurrence(piece, excel_row, col, "multi"))

        for col in single_cols:
            val = row[col]
            if isinstance(val, str) and val.strip():
                occurrences.append(NameOccurrence(val.strip(), excel_row, col, "single"))

    return occurrences


def build_clusters(unique_names: list[str]) -> list[list[str]]:
    """
    Group names by fuzzy similarity. Two names land in the same cluster when their
    normalized forms are >= SIMILARITY_THRESHOLD similar.
    """
    norm = {n: normalize(n) for n in unique_names}
    parent = {n: n for n in unique_names}

    def find(x: str) -> str:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    sorted_names = sorted(unique_names, key=lambda n: norm[n])
    for i, a in enumerate(sorted_names):
        if len(norm[a]) < MIN_NAME_LENGTH:
            continue
        for b in sorted_names[i + 1 :]:
            if len(norm[b]) < MIN_NAME_LENGTH:
                continue
            # Fast prefix prune: if first letter differs and edit distance is large, skip.
            if abs(len(norm[a]) - len(norm[b])) > 6:
                continue
            score = fuzz.ratio(norm[a], norm[b])
            if score >= SIMILARITY_THRESHOLD:
                union(a, b)

    clusters: dict[str, list[str]] = defaultdict(list)
    for n in unique_names:
        clusters[find(n)].append(n)
    return [c for c in clusters.values() if len(c) > 1]


def pick_canonical(cluster: list[str], counts: dict[str, int]) -> str:
    """
    Choose the 'correct' spelling for a cluster:
      1. Highest occurrence count wins.
      2. Tie-breaker: prefer the one with accents preserved (longer normalized form).
      3. Final tie-breaker: alphabetical.
    """
    return sorted(
        cluster,
        key=lambda n: (-counts[n], -len(n), n.lower()),
    )[0]


# ---- Duplicate-ISRC check ---------------------------------------------------


def find_isrc_conflicts(
    df: pd.DataFrame,
    tracks_sheet: str,
    isrc_leave_records: list[frozenset] | None = None,
    prior_isrc_actions: dict[tuple[str, int], dict] | None = None,
    check_same_release: bool = True,
) -> tuple[list[dict], list[dict]]:
    """
    Flag two kinds of ISRC problems (applies to all formats):

      1. Same ISRC → different Track Title (after normalization).
         An ISRC uniquely identifies a recording; if two rows share the same
         ISRC but have different titles one of them is wrong.

      2. Same ISRC appearing more than once within the same release (same UPC).
         The same recording on multiple albums is fine; appearing twice on the
         same album is a duplicate row.

    The previous check (same ISRC + different artist) produced too many false
    positives because the same track legitimately appears under different
    credited artists across releases.

    Returns (issues, conflict_summary).
    """
    isrc_leave_records = isrc_leave_records or []
    prior_isrc_actions = prior_isrc_actions or {}

    if ISRC_COLUMN not in df.columns:
        return [], []

    has_title = TRACK_TITLE_COLUMN in df.columns
    has_artist = DISPLAY_ARTIST_COLUMN in df.columns
    has_upc = "UPC" in df.columns
    # Splits format uses "Release ID" instead of UPC as the release grouper.
    release_col = "UPC" if has_upc else ("Release ID" if "Release ID" in df.columns else None)

    grouped: dict[str, list[dict]] = defaultdict(list)
    for idx, row in df.iterrows():
        isrc_raw = row.get(ISRC_COLUMN)
        if not isinstance(isrc_raw, str) or not isrc_raw.strip():
            continue
        isrc = isrc_raw.strip()
        title = str(row.get(TRACK_TITLE_COLUMN, "") or "").strip() if has_title else ""
        artist = str(row.get(DISPLAY_ARTIST_COLUMN, "") or "").strip() if has_artist else ""
        release = str(row.get(release_col, "") or "").strip() if release_col else ""
        grouped[isrc].append({
            "row": int(idx) + 2,
            "isrc": isrc,
            "title": title,
            "artist": artist,
            "release": release,
        })

    issues: list[dict] = []
    per_row_summary: list[dict] = []
    conflict_id = 1

    for isrc in sorted(grouped.keys()):
        occs = grouped[isrc]
        if len(occs) < 2:
            continue

        # ---- Check 1: same ISRC → different Track Title ----------------------
        unique_titles_norm = {normalize(o["title"]) for o in occs if o["title"]}
        title_conflict = len(unique_titles_norm) > 1

        # ---- Check 2: same ISRC + same release appearing more than once ------
        # Skipped for splits-correction sheets where multiple rows per ISRC/
        # release is the expected data structure (one row per allocatee).
        release_conflict = False
        if check_same_release and release_col:
            release_counts: dict[str, int] = defaultdict(int)
            for o in occs:
                if o["release"]:
                    release_counts[o["release"]] += 1
            release_conflict = any(n > 1 for n in release_counts.values())

        if not title_conflict and not release_conflict:
            continue

        if _isrc_conflict_is_left(occs, isrc_leave_records):
            continue

        seen_titles: list[str] = []
        for o in occs:
            if o["title"] and o["title"] not in seen_titles:
                seen_titles.append(o["title"])

        rows_str = ", ".join(str(o["row"]) for o in occs)
        cluster_label = f"I{conflict_id}"

        for o in occs:
            reasons = []
            if title_conflict:
                reasons.append(f"ISRC maps to multiple track titles: {'; '.join(seen_titles)}.")
            if release_conflict and o["release"] and release_counts.get(o["release"], 0) > 1:
                reasons.append(f"ISRC appears more than once in release {o['release']}.")
            if not reasons:
                continue
            issues.append({
                "Type": TYPE_ISRC,
                "Sheet": tracks_sheet,
                "Excel Row": o["row"],
                "Column": ISRC_COLUMN,
                "Found Value": isrc,
                "Suggested Value": "",
                "Similarity": "",
                "Cluster": cluster_label,
                "Notes": f"Rows {rows_str}. " + " ".join(reasons),
            })

            prior = prior_isrc_actions.get((isrc, o["row"]), {})
            per_row_summary.append({
                "Conflict": cluster_label,
                "ISRC": isrc,
                "Excel Row": o["row"],
                "Track Title": o["title"],
                "Track Display Artist": o["artist"],
                "Confirm OK?": prior.get("confirm_ok", ""),
                "Corrected ISRC": prior.get("corrected_isrc", ""),
            })

        conflict_id += 1

    issues.sort(key=lambda r: (r["Cluster"], r["Excel Row"]))
    per_row_summary.sort(key=lambda r: (
        int(r["Conflict"][1:]) if r["Conflict"][1:].isdigit() else 0,
        r["Excel Row"],
    ))
    return issues, per_row_summary


# ---- Missing-required-fields check ------------------------------------------


def _missing_kind(value) -> str | None:
    """
    Return None if the value is acceptably filled; otherwise a short label
    describing why it's considered missing ("blank", "whitespace only", or
    "placeholder").
    """
    if value is None:
        return "blank"
    # NaN check (pandas reads empty cells as float NaN)
    try:
        if isinstance(value, float) and value != value:  # NaN != NaN
            return "blank"
    except Exception:
        pass

    s = str(value)
    stripped = s.strip()
    if stripped == "":
        return "blank" if s == "" else "whitespace only"
    if stripped.lower() in PLACEHOLDER_VALUES:
        return f"placeholder ('{stripped}')"
    return None


# ---- Auto-suggest helpers for the missing-fields fix workflow -------------

# Fields that travel together at the release level — i.e., normally identical
# for every row sharing the same UPC. When one row is missing one of these
# and the other rows in the same release agree on a single value, we suggest
# that value as the fill.
RELEASE_LEVEL_FIELDS = {
    "UPC",
    "Release Title",
    "Release Artist",
    "Release Date",
    "Release Type",
    "Track P Line",
    "Track C Line",
}

# Fields where, if ALL other non-empty rows in the sheet agree on one value,
# we suggest that value (typical for label-wide constants).
SHEET_LEVEL_FIELDS = {
    "Track Language",
}


def _is_filled(v) -> bool:
    """True if v is a real value (not None, NaN, blank, whitespace, or a placeholder)."""
    return _missing_kind(v) is None


def _stringify(v) -> str:
    """Render a cell value as a clean string for use as a Fill Value suggestion."""
    if v is None:
        return ""
    # pandas may give us a float for an integer column (e.g., UPC 1.9e11). Clean those up.
    if isinstance(v, float):
        if pd.isna(v):
            return ""
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def _suggest_fill(
    col: str,
    row_idx: int,
    df: pd.DataFrame,
    upc_col: str | None,
    artist1_track_col: str | None,
) -> tuple[str, str]:
    """
    Return (suggested_value, reason). Empty strings if no confident suggestion.

    Suggestion rules:
      • Track Display Artist → copy from Artist 1 Name on Track on the same row.
      • Release-level field   → if all other rows with the same UPC agree on a
                                value, suggest it.
      • Sheet-level field     → if all non-empty rows in the sheet agree on a
                                value, suggest it.
      • Anything else         → no suggestion.
    """
    # 1. Track Display Artist → Artist 1 Name on Track (or its Mau5trap variant).
    if col == DISPLAY_ARTIST_COLUMN and artist1_track_col and artist1_track_col in df.columns:
        v = df.at[row_idx, artist1_track_col] if row_idx in df.index else None
        if _is_filled(v):
            return _stringify(v), f"copied from {artist1_track_col} on this row"

    # 2. Release-level: look at other rows with the same UPC.
    if col in RELEASE_LEVEL_FIELDS and upc_col and upc_col in df.columns and col != upc_col:
        upc_val = df.at[row_idx, upc_col] if row_idx in df.index else None
        if _is_filled(upc_val):
            same_release = df[df[upc_col].astype(str).str.strip() == _stringify(upc_val)]
            other_rows = same_release.drop(index=row_idx, errors="ignore")
            values = {_stringify(v) for v in other_rows[col] if _is_filled(v)}
            if len(values) == 1:
                v = next(iter(values))
                return v, f"consistent across other rows with UPC {_stringify(upc_val)}"

    # 3. Sheet-level: look at all rows.
    if col in SHEET_LEVEL_FIELDS and col in df.columns:
        values = {_stringify(v) for v in df[col] if _is_filled(v)}
        if len(values) == 1:
            v = next(iter(values))
            return v, f"every other row in the sheet uses {v!r}"

    return "", ""


def load_prior_format_corrections(issues_path: Path | None) -> dict[tuple[int, str], str]:
    """
    Read the Format Corrections tab from a previous _issues.xlsx so a re-scan
    doesn't lose the user's typed Corrected Values for ISRC/UPC format errors.
    Returns a map keyed by (Excel Row, Column).
    """
    if not issues_path or not issues_path.exists():
        return {}
    try:
        wb = load_workbook(issues_path, read_only=True, data_only=True)
        if "Format Corrections" not in wb.sheetnames:
            return {}
        ws = wb["Format Corrections"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        header = [str(h or "").strip() for h in rows[0]]
        if "Excel Row" not in header or "Column" not in header or "Corrected Value" not in header:
            return {}
        i_row = header.index("Excel Row")
        i_col = header.index("Column")
        i_corr = header.index("Corrected Value")
        out: dict[tuple[int, str], str] = {}
        for r in rows[1:]:
            if not r or i_row >= len(r) or i_col >= len(r):
                continue
            try:
                row_num = int(r[i_row])
            except (TypeError, ValueError):
                continue
            colname = str(r[i_col] or "").strip()
            if not colname:
                continue
            v = r[i_corr] if 0 <= i_corr < len(r) else ""
            out[(row_num, colname)] = "" if v is None else str(v).strip()
        return out
    except Exception:
        return {}


def load_prior_splits_review(issues_path: Path | None) -> dict[tuple[int, str], str]:
    """
    Read the Master Splits Review tab from a previous _issues.xlsx so a re-scan
    doesn't lose the user's edited split values. Returns a map keyed by
    (Excel Row, split_column_name).
    """
    if not issues_path or not issues_path.exists():
        return {}
    try:
        wb = load_workbook(issues_path, read_only=True, data_only=True)
        if "Master Splits Review" not in wb.sheetnames:
            return {}
        ws = wb["Master Splits Review"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        header = [str(h or "").strip() for h in rows[0]]
        if "Excel Row" not in header:
            return {}
        i_row = header.index("Excel Row")
        # Split columns are anything that starts with "Artist " and ends with
        # " Master Split". Indices in the header.
        split_col_idxs = [
            (j, h) for j, h in enumerate(header)
            if h.startswith("Artist ") and h.endswith(" Master Split")
        ]
        out: dict[tuple[int, str], str] = {}
        for r in rows[1:]:
            if not r or i_row >= len(r):
                continue
            try:
                row_num = int(r[i_row])
            except (TypeError, ValueError):
                continue
            for j, name in split_col_idxs:
                if j >= len(r):
                    continue
                v = r[j]
                out[(row_num, name)] = "" if v is None else str(v).strip()
        return out
    except Exception:
        return {}


def load_prior_missing_actions(issues_path: Path | None) -> dict[tuple[int, str], str]:
    """
    Read the Missing Field Corrections tab from a previous _issues.xlsx so a
    re-scan doesn't lose the user's typed Fill Values. Returns a map keyed by
    (Excel Row, Column).
    """
    if not issues_path or not issues_path.exists():
        return {}
    try:
        wb = load_workbook(issues_path, read_only=True, data_only=True)
        if "Missing Field Corrections" not in wb.sheetnames:
            return {}
        ws = wb["Missing Field Corrections"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        header = [str(h or "").strip() for h in rows[0]]
        if "Excel Row" not in header or "Column" not in header or "Fill Value" not in header:
            return {}
        i_row = header.index("Excel Row")
        i_col = header.index("Column")
        i_fill = header.index("Fill Value")
        out: dict[tuple[int, str], str] = {}
        for r in rows[1:]:
            if not r or i_row >= len(r) or i_col >= len(r):
                continue
            try:
                row_num = int(r[i_row])
            except (TypeError, ValueError):
                continue
            colname = str(r[i_col] or "").strip()
            if not colname:
                continue
            fill_val = r[i_fill] if 0 <= i_fill < len(r) else ""
            out[(row_num, colname)] = "" if fill_val is None else str(fill_val).strip()
        return out
    except Exception:
        return {}


def find_missing_required(
    df: pd.DataFrame,
    tracks_sheet: str,
    prior_missing_actions: dict[tuple[int, str], str] | None = None,
    required_fields_override: list | None = None,
) -> tuple[list[dict], list[dict], list[dict]]:
    """
    For each required field that exists in the sheet, flag every row whose value
    is missing (blank, whitespace-only, or a placeholder like 'N/A' or 'TBD').
    Tuple entries in REQUIRED_FIELDS are "any of these candidates" — the scanner
    uses the first present candidate.

    Returns (issues, missing_summary, per_cell_summary):
      • issues          — one dict per flagged cell (for the Issues tab + annotated copy)
      • missing_summary — one dict per required column (for the existing Missing Fields tab)
      • per_cell_summary — one dict per flagged cell (for the new Missing Field Corrections
                           tab, where the user types a Fill Value to apply in bulk)
    """
    prior_missing_actions = prior_missing_actions or {}

    fields_to_check = required_fields_override if required_fields_override is not None else REQUIRED_FIELDS

    # Resolve each requirement to either an actual column name (if present) or None.
    resolved: list[tuple[object, str | None]] = []
    for req in fields_to_check:
        resolved.append((req, required_field_present(req, df.columns)))

    present_cols = [c for _, c in resolved if c]
    counts: dict[str, dict[str, int]] = {
        c: {"blank": 0, "whitespace only": 0, "placeholder": 0} for c in present_cols
    }
    issues: list[dict] = []
    per_cell_summary: list[dict] = []

    # Resolve the columns we need for auto-suggest (UPC + Artist 1 Name on Track).
    upc_col = "UPC" if "UPC" in df.columns else None
    artist1_track_col = None
    for cand in ("Artist 1 Name on Track", "Artist 1 Client Name On Track"):
        if cand in df.columns:
            artist1_track_col = cand
            break

    for idx, row in df.iterrows():
        excel_row = int(idx) + 2
        for col in present_cols:
            kind = _missing_kind(row.get(col))
            if kind is None:
                continue
            issues.append({
                "Type": TYPE_MISSING,
                "Sheet": tracks_sheet,
                "Excel Row": excel_row,
                "Column": col,
                "Found Value": "" if kind == "blank" else str(row.get(col)),
                "Suggested Value": "(required — fill this in)",
                "Similarity": "",
                "Cluster": f"M-{col}",
                "Notes": kind,
            })
            bucket = "placeholder" if kind.startswith("placeholder") else kind
            counts[col][bucket] = counts[col].get(bucket, 0) + 1

            suggested, reason = _suggest_fill(col, idx, df, upc_col, artist1_track_col)
            # Carry the user's prior typed Fill Value forward (verbatim, including
            # empty if they explicitly cleared it). Otherwise pre-fill with the
            # suggestion so the most common case is one-glance confirm.
            prior = prior_missing_actions.get((excel_row, col))
            if prior is None:
                fill_value = suggested  # may be empty if no suggestion
            else:
                fill_value = prior
            track_title = _stringify(row.get(TRACK_TITLE_COLUMN, "")) if TRACK_TITLE_COLUMN in df.columns else ""
            display_artist = _stringify(row.get(DISPLAY_ARTIST_COLUMN, "")) if DISPLAY_ARTIST_COLUMN in df.columns else ""
            per_cell_summary.append({
                "Excel Row": excel_row,
                "Column": col,
                "Track Title": track_title,
                "Track Display Artist": display_artist,
                "Reason for missing": kind,
                "Suggested Fill": suggested,
                "Fill Value": fill_value,
                "Suggestion source": reason,
            })

    missing_summary: list[dict] = []
    for col in present_cols:
        c = counts[col]
        total = c["blank"] + c["whitespace only"] + c["placeholder"]
        if total == 0:
            continue
        breakdown_parts = []
        if c["blank"]:
            breakdown_parts.append(f"{c['blank']} blank")
        if c["whitespace only"]:
            breakdown_parts.append(f"{c['whitespace only']} whitespace-only")
        if c["placeholder"]:
            breakdown_parts.append(f"{c['placeholder']} placeholder")
        missing_summary.append({
            "Required field": col,
            "Rows missing": total,
            "Breakdown": ", ".join(breakdown_parts),
        })

    # Note any requirements whose candidates are all absent from the sheet — likely
    # a layout the scanner doesn't fully recognize.
    for req, present in resolved:
        if present is None:
            label = req if isinstance(req, str) else " / ".join(req)
            missing_summary.append({
                "Required field": label,
                "Rows missing": "—",
                "Breakdown": "no matching column found in sheet",
            })

    missing_summary.sort(
        key=lambda r: (-(r["Rows missing"] if isinstance(r["Rows missing"], int) else -1),
                       r["Required field"])
    )
    issues.sort(key=lambda r: (r["Column"], r["Excel Row"]))
    per_cell_summary.sort(key=lambda r: (r["Column"], r["Excel Row"]))
    return issues, missing_summary, per_cell_summary


# ---- Format-validation check ------------------------------------------------


def _has_pct_problem(cell) -> bool:
    """True if this cell exhibits the percent-stored-as-decimal problem."""
    fmt = cell.number_format or ""
    if "%" in fmt:
        return True
    if isinstance(cell.value, str) and cell.value.strip().endswith("%"):
        return True
    return False


def _is_empty_cell(cell) -> bool:
    v = cell.value
    if v is None:
        return True
    if isinstance(v, float) and v != v:  # NaN
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _split_value_as_percent(cell, column_is_pct: bool) -> float | None:
    """
    Convert a master-split cell to its 'displayed percent' value (e.g., 25 for 25%).
    Returns None if the cell isn't a usable number.
    """
    v = cell.value
    if v is None:
        return None
    fmt = cell.number_format or ""
    try:
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            had_pct = s.endswith("%")
            if had_pct:
                s = s.rstrip("%").strip()
            n = float(s)
            # Strings with explicit '%' are already in displayed-percent form.
            return n
        n = float(v)
    except (TypeError, ValueError):
        return None
    # If the number_format applies % (column-wide or this cell), values < 1 are
    # the decimal form (0.25 = 25%). Multiply by 100 to get the displayed percent.
    if (column_is_pct or "%" in fmt) and 0 < n <= 1:
        n *= 100
    return n


class _CellInfo:
    """Tiny stand-in for an openpyxl Cell, with just .value and .number_format.
    Built from a single iter_rows pass so we don't need slow random-access ws.cell()
    calls (which can take 10s+ on big workbooks)."""
    __slots__ = ("value", "number_format")

    def __init__(self, value=None, number_format: str = ""):
        self.value = value
        self.number_format = number_format

    def __repr__(self) -> str:
        return f"_CellInfo({self.value!r}, {self.number_format!r})"


def find_format_issues(
    input_path: Path,
    df: pd.DataFrame,
    tracks_sheet: str,
    prior_format_corrections: dict[tuple[int, str], str] | None = None,
    prior_splits_review: dict[tuple[int, str], str] | None = None,
) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict]]:
    """
    Run all format-related checks. Returns:
        issues               — list of issue rows (Type=Format issue) for the Issues tab
                               and the annotated copy.
        column_summary       — one row per column-wide format problem (for Format Issues tab)
        per_row_summary      — one row per row-level format failure (e.g., split sum != 100)
        format_corrections   — one row per offending CELL for ISRC/UPC format errors
                               (used to render the Format Corrections review tab where
                               the user types Corrected Value).
        splits_review        — one row per offending row for splits-not-summing-to-100,
                               with each Artist N Master Split value laid out as its own
                               column (used to render the Master Splits Review tab).
    """
    prior_format_corrections = prior_format_corrections or {}
    prior_splits_review = prior_splits_review or {}
    # Read-only mode is dramatically faster on big multi-sheet workbooks (e.g.,
    # 4× faster on 17-tab files). We only need values and number_formats here.
    wb = load_workbook(input_path, read_only=True, data_only=False)
    try:
        ws = wb[tracks_sheet]

        # ---- Single iter_rows pass: header + cell cache for needed cols ----
        header_to_col: dict[str, int] = {}
        cell_cache: dict[tuple[int, int], _CellInfo] = {}
        max_row = 1
        rows_iter = ws.iter_rows(values_only=False)

        first_row = next(rows_iter, None)
        if first_row is not None:
            for cell in first_row:
                if cell.value is not None:
                    header_to_col[str(cell.value)] = cell.column

        # Determine which columns we need cached data for.
        split_col_indices: dict[str, int] = {}
        for n in range(1, MAX_ARTIST_SLOTS + 1):
            col_name = f"Artist {n} Master Split"
            if col_name in header_to_col:
                split_col_indices[col_name] = header_to_col[col_name]

        upc_col_idx = header_to_col.get(UPC_COLUMN)
        isrc_col_idx = header_to_col.get(ISRC_COLUMN)
        needed_cols: set[int] = set(split_col_indices.values())
        if upc_col_idx is not None:
            needed_cols.add(upc_col_idx)
        if isrc_col_idx is not None:
            needed_cols.add(isrc_col_idx)

        for row in rows_iter:
            for cell in row:
                if not hasattr(cell, "column") or cell.column is None:
                    continue
                if cell.column in needed_cols:
                    cell_cache[(cell.row, cell.column)] = _CellInfo(
                        cell.value, cell.number_format or ""
                    )
                    if cell.row > max_row:
                        max_row = cell.row
    finally:
        wb.close()

    def _cell_at(r: int, col_idx: int) -> _CellInfo:
        return cell_cache.get((r, col_idx)) or _CellInfo(None, "")

    issues: list[dict] = []
    column_summary: list[dict] = []
    per_row_summary: list[dict] = []
    format_corrections: list[dict] = []
    splits_review: list[dict] = []

    # Tiny helpers — pulled into closures so we can use the dataframe + caches.
    def _track_title_for(r: int) -> str:
        if TRACK_TITLE_COLUMN not in df.columns:
            return ""
        idx = r - 2  # Excel row → 0-based dataframe index
        if idx not in df.index:
            return ""
        v = df.at[idx, TRACK_TITLE_COLUMN]
        return _stringify(v) if v is not None else ""

    def _display_artist_for(r: int) -> str:
        if DISPLAY_ARTIST_COLUMN not in df.columns:
            return ""
        idx = r - 2
        if idx not in df.index:
            return ""
        v = df.at[idx, DISPLAY_ARTIST_COLUMN]
        return _stringify(v) if v is not None else ""

    # ---- Master-split % formatting -----------------------------------------
    column_wide_pct: set[str] = set()
    for col_name, col_idx in split_col_indices.items():
        non_empty_rows: list[int] = []
        problem_rows: list[int] = []
        for r in range(2, max_row + 1):
            cell = _cell_at(r, col_idx)
            if _is_empty_cell(cell):
                continue
            non_empty_rows.append(r)
            if _has_pct_problem(cell):
                problem_rows.append(r)

        if not non_empty_rows or not problem_rows:
            continue

        ratio = len(problem_rows) / len(non_empty_rows)
        if ratio >= SPLIT_PCT_COLUMN_THRESHOLD:
            # Column-wide problem — flag the header cell, skip per-cell flags.
            column_wide_pct.add(col_name)
            issues.append({
                "Type": TYPE_FORMAT,
                "Sheet": tracks_sheet,
                "Excel Row": 1,
                "Column": col_name,
                "Found Value": "(applies to entire column)",
                "Suggested Value": "Plain integer percentages (25, not 0.25 with % format)",
                "Similarity": "",
                "Cluster": f"F-PCT-{col_name}",
                "Notes": (
                    f"This column is formatted as a percentage. Cells display 25% "
                    f"but Excel actually stores 0.25 — when uploaded, the raw 0.25 is read "
                    f"instead of 25. Affects {len(problem_rows)} of {len(non_empty_rows)} "
                    f"non-empty cells."
                ),
            })
            column_summary.append({
                "Issue": "Master split column with % formatting",
                "Column": col_name,
                "Rows affected": f"{len(problem_rows)} of {len(non_empty_rows)} non-empty",
                "Why it matters": (
                    "Cells store the decimal form (0.25). Uploading reads 0.25 instead of 25."
                ),
            })
        else:
            # Occasional problem — flag each cell individually.
            for r in problem_rows:
                cell = _cell_at(r, col_idx)
                shown = str(cell.value) if cell.value is not None else ""
                fmt_note = f" (format: {cell.number_format})" if "%" in cell.number_format else ""
                issues.append({
                    "Type": TYPE_FORMAT,
                    "Sheet": tracks_sheet,
                    "Excel Row": r,
                    "Column": col_name,
                    "Found Value": shown + fmt_note,
                    "Suggested Value": "Plain integer percentage (e.g., 25, not 0.25 with % format)",
                    "Similarity": "",
                    "Cluster": f"F-PCT-{col_name}",
                    "Notes": (
                        "Cell uses % formatting — stored value is decimal (0.25 displays as 25%); "
                        "ingestion will read the raw 0.25."
                    ),
                })

    # ---- Master-split sum to 100 -------------------------------------------
    if split_col_indices:
        for r in range(2, max_row + 1):
            total = 0.0
            any_value = False
            for col_name, col_idx in split_col_indices.items():
                cell = _cell_at(r, col_idx)
                if _is_empty_cell(cell):
                    continue
                pct = _split_value_as_percent(cell, col_name in column_wide_pct)
                if pct is None:
                    continue
                total += pct
                any_value = True
            if not any_value:
                continue
            if abs(total - SPLIT_SUM_TARGET) > SPLIT_SUM_TOLERANCE:
                a1_col_name = (
                    "Artist 1 Master Split"
                    if "Artist 1 Master Split" in split_col_indices
                    else next(iter(split_col_indices.keys()))
                )
                issues.append({
                    "Type": TYPE_FORMAT,
                    "Sheet": tracks_sheet,
                    "Excel Row": r,
                    "Column": a1_col_name,
                    "Found Value": f"row total: {total:.2f}",
                    "Suggested Value": f"{SPLIT_SUM_TARGET:.0f}",
                    "Similarity": "",
                    "Cluster": "F-SPLIT-SUM",
                    "Notes": (
                        f"Master splits in this row sum to {total:.2f}%, not 100%. "
                        f"Difference: {total - SPLIT_SUM_TARGET:+.2f}."
                    ),
                })
                per_row_summary.append({
                    "Issue": "Master splits don't sum to 100",
                    "Excel Row": r,
                    "Total %": f"{total:.2f}",
                    "Difference": f"{total - SPLIT_SUM_TARGET:+.2f}",
                })
                # Build a per-row review entry with each Artist N Master Split
                # value laid out as its own field. The user edits the values
                # in-line in the Master Splits Review tab.
                review_row: dict = {
                    "Excel Row": r,
                    "Track Title": _track_title_for(r),
                    "Track Display Artist": _display_artist_for(r),
                }
                for col_name, col_idx in split_col_indices.items():
                    cell = _cell_at(r, col_idx)
                    pct = _split_value_as_percent(cell, col_name in column_wide_pct)
                    # Pre-fill with the displayed percentage (e.g., 25, not 0.25).
                    # If the user edited this cell in a previous run, lift their value.
                    prior_val = prior_splits_review.get((r, col_name), None)
                    if prior_val is not None:
                        review_row[col_name] = prior_val
                    elif pct is None:
                        review_row[col_name] = ""
                    else:
                        # Keep ints as ints (50 not 50.0); show floats with up to 2 dp.
                        review_row[col_name] = (
                            int(pct) if float(pct).is_integer() else round(pct, 2)
                        )
                review_row["Current Sum"] = f"{total:.2f}"
                splits_review.append(review_row)

    # ---- ISRC format -------------------------------------------------------
    if isrc_col_idx is not None:
        bad = 0
        for r in range(2, max_row + 1):
            cell = _cell_at(r, isrc_col_idx)
            if _is_empty_cell(cell):
                continue
            raw = str(cell.value).strip()
            cleaned = raw.upper().replace("-", "").replace(" ", "")
            if not ISRC_PATTERN.match(cleaned):
                bad += 1
                issues.append({
                    "Type": TYPE_FORMAT,
                    "Sheet": tracks_sheet,
                    "Excel Row": r,
                    "Column": ISRC_COLUMN,
                    "Found Value": raw,
                    "Suggested Value": "12 chars: 2-letter country + 3 alphanumeric registrant + 7 digits (year + designation)",
                    "Similarity": "",
                    "Cluster": "F-ISRC",
                    "Notes": "Doesn't match ISRC format (CCRRRYYNNNNN, 12 chars).",
                })
                prior = prior_format_corrections.get((r, ISRC_COLUMN), None)
                format_corrections.append({
                    "Type": "ISRC format",
                    "Excel Row": r,
                    "Column": ISRC_COLUMN,
                    "Track Title": _track_title_for(r),
                    "Track Display Artist": _display_artist_for(r),
                    "Found Value": raw,
                    "Corrected Value": prior if prior is not None else "",
                    "Notes": "Must be 12 chars: 2-letter country + 3 alphanumeric registrant + 7 digits.",
                })
        if bad:
            per_row_summary.append({
                "Issue": "Malformed ISRC values",
                "Excel Row": "—",
                "Total %": "—",
                "Difference": f"{bad} cells",
            })

    # ---- UPC format --------------------------------------------------------
    if upc_col_idx is not None:
        bad = 0
        for r in range(2, max_row + 1):
            cell = _cell_at(r, upc_col_idx)
            if _is_empty_cell(cell):
                continue
            v = cell.value
            if isinstance(v, float) and v.is_integer():
                upc_str = str(int(v))
            else:
                upc_str = str(v).strip()
            upc_str_cleaned = upc_str.replace("-", "").replace(" ", "")
            if not UPC_PATTERN.match(upc_str_cleaned):
                bad += 1
                issues.append({
                    "Type": TYPE_FORMAT,
                    "Sheet": tracks_sheet,
                    "Excel Row": r,
                    "Column": UPC_COLUMN,
                    "Found Value": upc_str,
                    "Suggested Value": "12 or 13 digits, stored as text",
                    "Similarity": "",
                    "Cluster": "F-UPC",
                    "Notes": "UPC isn't 12–13 digits (or contains non-digit characters).",
                })
                prior = prior_format_corrections.get((r, UPC_COLUMN), None)
                format_corrections.append({
                    "Type": "UPC format",
                    "Excel Row": r,
                    "Column": UPC_COLUMN,
                    "Track Title": _track_title_for(r),
                    "Track Display Artist": _display_artist_for(r),
                    "Found Value": upc_str,
                    "Corrected Value": prior if prior is not None else "",
                    "Notes": "Must be 12 or 13 digits.",
                })
        if bad:
            per_row_summary.append({
                "Issue": "Malformed UPC values",
                "Excel Row": "—",
                "Total %": "—",
                "Difference": f"{bad} cells",
            })

    issues.sort(key=lambda r: (r["Cluster"], r["Excel Row"]))
    format_corrections.sort(key=lambda r: (r["Type"], r["Excel Row"]))
    splits_review.sort(key=lambda r: r["Excel Row"])
    return issues, column_summary, per_row_summary, format_corrections, splits_review


# ---- Report writer ----------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
HIGH_FILL = PatternFill("solid", start_color="FCE4D6")  # light red for the "Found Value"
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")  # soft yellow — the user types here


def autosize(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_width, min(longest + 2, max_width)
        )


def write_report(
    output_path: Path,
    issues: list[dict],
    cluster_summary: list[dict],
    isrc_summary: list[dict],
    missing_summary: list[dict],
    missing_per_cell: list[dict],
    format_column_summary: list[dict],
    format_row_summary: list[dict],
    format_corrections: list[dict],
    splits_review: list[dict],
    source_path: Path,
    split_errors: list[dict] | None = None,
    id_mismatches: list[dict] | None = None,
    detected_format: str = "",
) -> None:
    wb = Workbook()

    # Issues sheet ----------------------------------------------------------
    ws = wb.active
    ws.title = "Issues"
    headers = [
        "Type",
        "Sheet",
        "Excel Row",
        "Column",
        "Found Value",
        "Suggested Value",
        "Similarity",
        "Cluster",
        "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    found_value_col = headers.index("Found Value") + 1
    for i, issue in enumerate(issues, start=2):
        ws.append([issue.get(h, "") for h in headers])
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_FILL
        ws.cell(row=i, column=found_value_col).fill = HIGH_FILL

    autosize(ws)

    # Artist Clusters sheet -------------------------------------------------
    # The "Correction" column is the user's input for the typo-correction
    # workflow. It's pre-filled with the canonical (most-frequent) variant so
    # most clusters can be confirmed at a glance — the user only has to edit
    # mistakes or clear cells they want skipped. Typing "LEAVE" marks a
    # cluster as intentionally distinct so future scans stop flagging it
    # (e.g., the DeBarge brothers).
    ws2 = wb.create_sheet("Artist Clusters")
    headers2 = ["Cluster", "Canonical", "Correction", "Variants (count)", "Total occurrences"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws2.freeze_panes = "A2"

    correction_col_idx = headers2.index("Correction") + 1
    for i, c in enumerate(cluster_summary, start=2):
        ws2.append([c.get(h, "") for h in headers2])
        if i % 2 == 0:
            for cell in ws2[i]:
                cell.fill = ALT_FILL
        # Highlight the editable column so it's visually obvious where to type.
        ws2.cell(row=i, column=correction_col_idx).fill = INPUT_FILL

    autosize(ws2, max_width=80)

    # ISRC Conflicts sheet --------------------------------------------------
    # One row per offending row (grouped visually by Conflict ID). The two
    # editable columns — "Confirm OK?" and "Corrected ISRC" — are highlighted
    # so it's obvious where to type. Workflow:
    #   • Type any of {OK, yes, x, ✓, true, 1} in "Confirm OK?" to mark this
    #     row as an intentional duplicate (persists across rescans).
    #   • Or type a replacement ISRC in "Corrected ISRC" to swap that row's
    #     Track ISRC when you run "Apply ISRC Corrections.command".
    #   • Leave both blank to defer the decision.
    ws3 = wb.create_sheet("ISRC Conflicts")
    headers3 = [
        "Conflict",
        "ISRC",
        "Excel Row",
        "Track Title",
        "Track Display Artist",
        "Confirm OK?",
        "Corrected ISRC",
    ]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws3.freeze_panes = "A2"

    confirm_col_idx = headers3.index("Confirm OK?") + 1
    corrected_col_idx = headers3.index("Corrected ISRC") + 1
    for i, c in enumerate(isrc_summary, start=2):
        ws3.append([c.get(h, "") for h in headers3])
        if i % 2 == 0:
            for cell in ws3[i]:
                cell.fill = ALT_FILL
        # Highlight the editable columns so it's visually obvious where to type.
        ws3.cell(row=i, column=confirm_col_idx).fill = INPUT_FILL
        ws3.cell(row=i, column=corrected_col_idx).fill = INPUT_FILL

    autosize(ws3, max_width=80)

    # Missing Fields sheet --------------------------------------------------
    ws4 = wb.create_sheet("Missing Fields")
    headers4 = ["Required field", "Rows missing", "Breakdown"]
    ws4.append(headers4)
    for cell in ws4[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws4.freeze_panes = "A2"

    for i, c in enumerate(missing_summary, start=2):
        ws4.append([c[h] for h in headers4])
        if i % 2 == 0:
            for cell in ws4[i]:
                cell.fill = ALT_FILL

    autosize(ws4, max_width=80)

    # Missing Field Corrections sheet --------------------------------------
    # One row per missing cell (the per-row review surface for the
    # "Apply Missing Field Corrections" launcher).
    #   • Suggested Fill is filled in by the scanner where it can guess
    #     confidently (release-level fields with consistent siblings, etc.).
    #   • Fill Value is yellow-highlighted, your input. Pre-filled with the
    #     suggestion when one exists. Carries forward across rescans.
    #   • Empty Fill Value → no action (defer).
    ws_mfc = wb.create_sheet("Missing Field Corrections")
    headers_mfc = [
        "Excel Row",
        "Column",
        "Track Title",
        "Track Display Artist",
        "Reason for missing",
        "Suggested Fill",
        "Fill Value",
        "Suggestion source",
    ]
    ws_mfc.append(headers_mfc)
    for cell in ws_mfc[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_mfc.freeze_panes = "A2"

    fill_value_idx = headers_mfc.index("Fill Value") + 1
    for i, c in enumerate(missing_per_cell, start=2):
        ws_mfc.append([c.get(h, "") for h in headers_mfc])
        if i % 2 == 0:
            for cell in ws_mfc[i]:
                cell.fill = ALT_FILL
        ws_mfc.cell(row=i, column=fill_value_idx).fill = INPUT_FILL

    autosize(ws_mfc, max_width=80)

    # Format Issues sheet ---------------------------------------------------
    ws5 = wb.create_sheet("Format Issues")
    ws5.append(["Column-wide format issues"])
    ws5["A1"].font = Font(bold=True, size=12, name="Arial")

    headers5a = ["Issue", "Column", "Rows affected", "Why it matters"]
    ws5.append(headers5a)
    for cell in ws5[2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    next_row = 3
    for c in format_column_summary:
        ws5.append([c[h] for h in headers5a])
        if next_row % 2 == 0:
            for cell in ws5[next_row]:
                cell.fill = ALT_FILL
        next_row += 1

    if not format_column_summary:
        ws5.append(["(none detected)"])
        next_row += 1

    next_row += 2
    ws5.cell(row=next_row, column=1, value="Row-level / per-cell format issues").font = Font(
        bold=True, size=12, name="Arial"
    )
    next_row += 1

    headers5b = ["Issue", "Excel Row", "Total %", "Difference"]
    for i, h in enumerate(headers5b, start=1):
        cell = ws5.cell(row=next_row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    next_row += 1

    if format_row_summary:
        for c in format_row_summary:
            for i, h in enumerate(headers5b, start=1):
                ws5.cell(row=next_row, column=i, value=c.get(h, ""))
            if next_row % 2 == 0:
                for cell in ws5[next_row]:
                    cell.fill = ALT_FILL
            next_row += 1
    else:
        ws5.cell(row=next_row, column=1, value="(none detected)")

    autosize(ws5, max_width=80)

    # Format Corrections sheet ---------------------------------------------
    # Per-cell review surface for ISRC + UPC format errors. The user types a
    # corrected value into the yellow column; the apply launcher writes it
    # to the (Excel Row, Column) cell on the annotated copy.
    ws_fc = wb.create_sheet("Format Corrections")
    headers_fc = [
        "Type",
        "Excel Row",
        "Column",
        "Track Title",
        "Track Display Artist",
        "Found Value",
        "Corrected Value",
        "Notes",
    ]
    ws_fc.append(headers_fc)
    for cell in ws_fc[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_fc.freeze_panes = "A2"

    corr_idx = headers_fc.index("Corrected Value") + 1
    for i, c in enumerate(format_corrections, start=2):
        ws_fc.append([c.get(h, "") for h in headers_fc])
        if i % 2 == 0:
            for cell in ws_fc[i]:
                cell.fill = ALT_FILL
        ws_fc.cell(row=i, column=corr_idx).fill = INPUT_FILL

    autosize(ws_fc, max_width=80)

    # Master Splits Review sheet -------------------------------------------
    # Per-row review surface for "splits don't sum to 100" rows. Each
    # Artist N Master Split is laid out as its own yellow editable column,
    # with the current values pre-filled. The user adjusts the values; the
    # apply launcher writes each non-empty cell back to the annotated copy.
    ws_msr = wb.create_sheet("Master Splits Review")
    # Discover which split column names are actually used in the data so we
    # only show the columns that exist in the sheet.
    split_cols_in_use: list[str] = []
    seen = set()
    for r in splits_review:
        for k in r.keys():
            if k.startswith("Artist ") and k.endswith(" Master Split") and k not in seen:
                split_cols_in_use.append(k)
                seen.add(k)
    # Sort by the artist number so the columns line up: A1, A2, A3, ...
    def _split_num(name: str) -> int:
        try:
            return int(name.split()[1])
        except (IndexError, ValueError):
            return 999
    split_cols_in_use.sort(key=_split_num)

    headers_msr = ["Excel Row", "Track Title", "Track Display Artist"] + split_cols_in_use + ["Current Sum"]
    ws_msr.append(headers_msr)
    for cell in ws_msr[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_msr.freeze_panes = "A2"

    split_first_idx = headers_msr.index(split_cols_in_use[0]) + 1 if split_cols_in_use else 0
    split_last_idx = headers_msr.index(split_cols_in_use[-1]) + 1 if split_cols_in_use else -1
    for i, c in enumerate(splits_review, start=2):
        ws_msr.append([c.get(h, "") for h in headers_msr])
        if i % 2 == 0:
            for cell in ws_msr[i]:
                cell.fill = ALT_FILL
        # Highlight every Artist N Master Split column on this row.
        if split_cols_in_use:
            for col_idx in range(split_first_idx, split_last_idx + 1):
                ws_msr.cell(row=i, column=col_idx).fill = INPUT_FILL

    autosize(ws_msr, max_width=80)

    # Split Errors tab (EpicWin splits format only) --------------------------
    if split_errors:
        ws_spl = wb.create_sheet("Split Errors")
        spl_headers = ["Release ID", "ISRC", "Track Title", "Rows", "Split Total", "Difference from 100"]
        ws_spl.append(spl_headers)
        for cell in ws_spl[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws_spl.freeze_panes = "A2"
        for i, row in enumerate(split_errors, start=2):
            ws_spl.append([row.get(h, "") for h in spl_headers])
            if i % 2 == 0:
                for cell in ws_spl[i]:
                    cell.fill = ALT_FILL
        autosize(ws_spl)

    # ID Mismatches tab (EpicWin splits format only) -------------------------
    if id_mismatches:
        ws_idm = wb.create_sheet("ID Mismatches")
        idm_headers = ["Type", "ID", "Names found", "Rows"]
        ws_idm.append(idm_headers)
        for cell in ws_idm[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws_idm.freeze_panes = "A2"
        for i, row in enumerate(id_mismatches, start=2):
            ws_idm.append([row.get(h, "") for h in idm_headers])
            if i % 2 == 0:
                for cell in ws_idm[i]:
                    cell.fill = ALT_FILL
        autosize(ws_idm)

    # About sheet -----------------------------------------------------------
    ws_about = wb.create_sheet("About", 0)
    n_artist = sum(1 for i in issues if i.get("Type") == TYPE_ARTIST)
    n_isrc = sum(1 for i in issues if i.get("Type") == TYPE_ISRC)
    n_missing = sum(1 for i in issues if i.get("Type") == TYPE_MISSING)
    n_format = sum(1 for i in issues if i.get("Type") == TYPE_FORMAT)
    n_splits = len(split_errors) if split_errors else 0
    n_id_mismatches = len(id_mismatches) if id_mismatches else 0
    info = [
        ["Metadata Sheet Scanner — Issues Report"],
        [],
        ["Source file", str(source_path)],
        ["Detected format", detected_format or "label-engine"],
        ["Checks run", "Artist-name typos · Duplicate ISRCs · Missing required fields · Format validation"],
        ["Artist similarity threshold", f"{SIMILARITY_THRESHOLD}%"],
        ["Min artist name length", str(MIN_NAME_LENGTH)],
        [],
        ["Artist-typo issues found", n_artist],
        ["Duplicate-ISRC issues found", n_isrc],
        ["Missing-field issues found", n_missing],
        ["Format issues found", n_format],
        ["Split errors found (EpicWin)", n_splits],
        ["ID mismatches found (EpicWin)", n_id_mismatches],
        ["Artist clusters identified", len(cluster_summary)],
        ["ISRC conflicts identified", len({r["Conflict"] for r in isrc_summary})],
        ["Column-wide format issues", len(format_column_summary)],
        [],
        ["How to read this:"],
        [
            "  • The Issues tab lists every flagged cell. Filter by 'Type' to switch between",
            "    artist typos, ISRC conflicts, missing required fields, and format issues.",
        ],
        [
            "  • The Artist Clusters tab shows each group of similar artist names with their",
            "    counts so you can verify the canonical pick before find-and-replace.",
        ],
        [
            "  • The ISRC Conflicts tab lists every offending row (one row per cell), grouped",
            "    by Conflict ID. Type any of {OK, yes, x, ✓} in 'Confirm OK?' to mark a row as",
            "    an intentional duplicate (persists across rescans), or type a replacement in",
            "    'Corrected ISRC' to swap that row's Track ISRC the next time you run the",
            "    'Apply ISRC Corrections' launcher.",
        ],
        [
            "  • The Missing Fields tab summarizes each required column and how many rows",
            "    are blank, whitespace-only, or filled with a placeholder ('N/A', 'TBD', etc.).",
        ],
        [
            "  • The Missing Field Corrections tab is the per-row review surface — one line",
            "    per missing cell, with a yellow Fill Value column you type into. Suggested Fill",
            "    is pre-populated where the scanner can guess (release-level fields with",
            "    consistent siblings, etc.). Run 'Apply Missing Field Corrections.command' to",
            "    write the values back into the annotated copy in bulk.",
        ],
        [
            "  • The Format Issues tab lists column-wide format problems (top section) and",
            "    row-level / per-cell ones (bottom). Includes ISRC format, UPC length, master",
            "    splits with %-sign formatting, and master splits not summing to 100%.",
        ],
        [
            "  • The Format Corrections tab is the per-cell review surface for ISRC + UPC",
            "    format errors — type the right value into the yellow Corrected Value column.",
        ],
        [
            "  • The Master Splits Review tab is the per-row review surface for splits not",
            "    summing to 100 — every Artist N Master Split is a yellow editable column.",
        ],
        [
            "  • Run 'Apply Format Corrections.command' to apply both tabs at once. The",
            "    launcher also auto-strips %-formatting from any column-wide-flagged columns.",
        ],
    ]
    for row in info:
        ws_about.append(row)
    ws_about["A1"].font = Font(bold=True, size=14, name="Arial")
    autosize(ws_about, max_width=100)

    wb.save(output_path)


# ---- Main -------------------------------------------------------------------


def analyze(
    input_path: Path,
    issues_output_path: Path | None = None,
    project_dir: Path | None = None,
) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], list[dict], list[dict], list[dict], list[dict], dict, list[dict], list[dict]]:
    """
    Run all checks on a sheet and return
    (issues, cluster_summary, isrc_summary, missing_summary, missing_per_cell,
     format_column_summary, format_row_summary, format_corrections,
     splits_review, stats).
    No file output — pure analysis. Used by both report and annotated modes.

    If issues_output_path is given and the file exists, any user-typed
    Correction values from a previous run are carried forward into the new
    cluster_summary so a re-scan in the middle of correction work doesn't
    blow away the user's progress.
    """
    project_root = Path(project_dir) if project_dir else Path(__file__).resolve().parent
    leave_records = load_leave_records(project_root)
    isrc_leave_records = load_isrc_leave_records(project_root)
    prior_corrections = load_prior_corrections(issues_output_path) if issues_output_path else []
    prior_isrc_actions = load_prior_isrc_actions(issues_output_path) if issues_output_path else {}
    prior_missing_actions = load_prior_missing_actions(issues_output_path) if issues_output_path else {}
    prior_format_corrections = load_prior_format_corrections(issues_output_path) if issues_output_path else {}
    prior_splits_review = load_prior_splits_review(issues_output_path) if issues_output_path else {}

    # Discover all sheets to scan and their schemas.
    sheets_to_scan = detect_all_sheets_to_scan(input_path)

    # If no schema matched, return immediately with zero issues and an "unknown"
    # format marker so the UI can surface a friendly message.
    if sheets_to_scan and sheets_to_scan[0][1].get("_key") == "unknown":
        first_sheet = sheets_to_scan[0][0]
        empty_stats = {
            "tracks_sheet": first_sheet,
            "detected_format": "unknown",
            "sheets_scanned": [first_sheet],
            "other_sheets_with_track_isrc": [],
            "occurrences": 0, "unique_names": 0, "artist_clusters": 0,
            "artist_typo_cells": 0, "isrc_conflicts": 0, "isrc_conflict_cells": 0,
            "missing_field_issues": 0, "format_issues": 0, "format_columns_wide": 0,
            "splits_errors": 0, "id_mismatches": 0, "splits_issues": 0, "total_issues": 0,
        }
        return ([], [], [], [], [], [], [], [], [], empty_stats, [], [])

    # Accumulate results across all sheets.
    all_occurrences: list = []
    all_isrc_issues: list[dict] = []
    all_isrc_summary: list[dict] = []
    all_missing_issues: list[dict] = []
    all_missing_summary: list[dict] = []
    all_missing_per_cell: list[dict] = []
    all_format_issues: list[dict] = []
    all_format_column_summary: list[dict] = []
    all_format_row_summary: list[dict] = []
    all_format_corrections: list[dict] = []
    all_splits_review: list[dict] = []
    all_split_errors: list[dict] = []
    all_id_mismatches: list[dict] = []
    detected_format_names: list[str] = []
    primary_sheet = sheets_to_scan[0][0] if sheets_to_scan else ""

    for sheet_name, schema in sheets_to_scan:
        schema_key = schema.get("_key", "label-engine")
        checks = schema.get("checks", set())
        required_override = schema.get("required_fields_override")
        detected_format_names.append(schema["display_name"])

        df_raw = pd.read_excel(input_path, sheet_name=sheet_name, header=0)
        df = normalize_columns(df_raw, schema)

        if "artist_typos" in checks:
            all_occurrences.extend(collect_occurrences(df))

        if "isrc_duplicates" in checks:
            isrc_iss, isrc_sum = find_isrc_conflicts(
                df, sheet_name,
                isrc_leave_records=isrc_leave_records,
                prior_isrc_actions=prior_isrc_actions,
                check_same_release=(schema_key != "epicwin-splits"),
            )
            all_isrc_issues.extend(isrc_iss)
            all_isrc_summary.extend(isrc_sum)

        if "missing_fields" in checks:
            miss_iss, miss_sum, miss_cell = find_missing_required(
                df, sheet_name,
                prior_missing_actions=prior_missing_actions,
                required_fields_override=required_override,
            )
            all_missing_issues.extend(miss_iss)
            all_missing_summary.extend(miss_sum)
            all_missing_per_cell.extend(miss_cell)

        if "format_validation" in checks:
            (fmt_iss, fmt_col, fmt_row, fmt_corr, spl_rev) = find_format_issues(
                input_path, df, sheet_name,
                prior_format_corrections=prior_format_corrections,
                prior_splits_review=prior_splits_review,
            )
            all_format_issues.extend(fmt_iss)
            all_format_column_summary.extend(fmt_col)
            all_format_row_summary.extend(fmt_row)
            all_format_corrections.extend(fmt_corr)
            all_splits_review.extend(spl_rev)

        if "splits_correction" in checks:
            spl_iss, spl_err, id_mis = find_splits_correction_issues(df, sheet_name)
            # Keep splits/ID issues separate from format_issues so the Formats
            # tab count only reflects genuine format validation errors.
            all_split_errors.extend(spl_err)
            all_id_mismatches.extend(id_mis)

    # ---- Artist-name clustering (across all sheets combined) ----------------
    unique_names = sorted({o.name for o in all_occurrences})
    counts: dict[str, int] = defaultdict(int)
    for o in all_occurrences:
        counts[o.name] += 1

    clusters = build_clusters(unique_names)
    if leave_records:
        clusters = [c for c in clusters if not _matches_leave(c, leave_records)]

    name_to_canonical: dict[str, tuple[str, int]] = {}
    cluster_summary: list[dict] = []
    for cid, cluster in enumerate(sorted(clusters, key=lambda c: -sum(counts[n] for n in c)), start=1):
        canonical = pick_canonical(cluster, counts)
        for n in cluster:
            name_to_canonical[n] = (canonical, cid)
        variants_str = "; ".join(
            f"{n} ({counts[n]})" for n in sorted(cluster, key=lambda x: -counts[x])
        )
        carried = _find_prior_correction(cluster, prior_corrections)
        correction_value = canonical if carried is None else carried
        cluster_summary.append({
            "Cluster": f"T{cid}",
            "Canonical": canonical,
            "Correction": correction_value,
            "Variants (count)": variants_str,
            "Total occurrences": sum(counts[n] for n in cluster),
        })

    artist_issues: list[dict] = []
    for o in all_occurrences:
        if o.name not in name_to_canonical:
            continue
        canonical, cid = name_to_canonical[o.name]
        if o.name == canonical:
            continue
        score = fuzz.ratio(normalize(o.name), normalize(canonical))
        notes = []
        if normalize(o.name) == normalize(canonical):
            notes.append("differs only by case/accents/whitespace")
        if o.name.strip() != o.name:
            notes.append("trailing/leading whitespace")
        if o.name.upper() == o.name and len(o.name) > 2:
            notes.append("ALL CAPS")
        artist_issues.append({
            "Type": TYPE_ARTIST,
            "Sheet": o.column,   # preserve sheet-column origin
            "Excel Row": o.row,
            "Column": o.column,
            "Found Value": o.name,
            "Suggested Value": canonical,
            "Similarity": score,
            "Cluster": f"T{cid}",
            "Notes": "; ".join(notes),
        })
    artist_issues.sort(key=lambda r: (r["Cluster"], r["Excel Row"]))

    # ---- Combine -----------------------------------------------------------
    # splits/ID issues are tracked separately and not mixed into format_issues
    # so the Formats tab count stays accurate. They appear in the downloaded
    # report's dedicated tabs and the UI's Splits tab.
    issues = artist_issues + all_isrc_issues + all_missing_issues + all_format_issues

    # Build format display string (deduplicated, preserving order).
    seen_fmt: set[str] = set()
    fmt_display_parts: list[str] = []
    for name in detected_format_names:
        if name not in seen_fmt:
            fmt_display_parts.append(name)
            seen_fmt.add(name)
    detected_format = " + ".join(fmt_display_parts) if fmt_display_parts else "label-engine"

    stats = {
        "tracks_sheet": primary_sheet,
        "detected_format": detected_format,
        "sheets_scanned": [s for s, _ in sheets_to_scan],
        "other_sheets_with_track_isrc": [],
        "occurrences": len(all_occurrences),
        "unique_names": len(unique_names),
        "artist_clusters": len(clusters),
        "artist_typo_cells": len(artist_issues),
        "isrc_conflicts": len({r["Conflict"] for r in all_isrc_summary}),
        "isrc_conflict_cells": len(all_isrc_issues),
        "missing_field_issues": len(all_missing_issues),
        "format_issues": len(all_format_issues),
        "format_columns_wide": len(all_format_column_summary),
        "splits_errors": len(all_split_errors),
        "id_mismatches": len(all_id_mismatches),
        "splits_issues": len(all_split_errors) + len(all_id_mismatches),
        "total_issues": len(issues) + len(all_split_errors) + len(all_id_mismatches),
    }
    return (
        issues,
        cluster_summary,
        all_isrc_summary,
        all_missing_summary,
        all_missing_per_cell,
        all_format_column_summary,
        all_format_row_summary,
        all_format_corrections,
        all_splits_review,
        stats,
        all_split_errors,
        all_id_mismatches,
    )


# ---- Annotated copy mode ----------------------------------------------------

ARTIST_FILL = PatternFill("solid", start_color="FCE4D6")    # light red — artist typos
ISRC_FILL = PatternFill("solid", start_color="FFF2CC")      # light yellow — ISRC conflicts
MISSING_FILL = PatternFill("solid", start_color="DDEBF7")   # light blue — missing required fields
FORMAT_FILL = PatternFill("solid", start_color="C6EFCE")    # light green — format issues
MIXED_FILL = PatternFill("solid", start_color="FFE699")     # darker yellow — multiple kinds in one cell
COMMENT_AUTHOR = "Metadata Decoder"

# Hex codes (without the alpha prefix Excel sometimes prepends) of the fills the
# decoder applies. Used to identify and strip prior decoder annotations on re-scans
# without touching anything the user added themselves.
# NOTE: "FCD5B4" is the *previous* format-issue color (light orange). It's kept here
# so re-scans of files annotated by older versions correctly clear the old highlights
# before applying the new green.
DECODER_FILL_HEXES = {"FCE4D6", "FFF2CC", "DDEBF7", "C6EFCE", "FCD5B4", "FFE699"}
DECODER_SUMMARY_SHEET = "Decoder Summary"


def _format_artist_issue(ci: dict) -> str:
    sim = ci.get("Similarity")
    sim_str = f"{float(sim):.0f}%" if isinstance(sim, (int, float)) else str(sim)
    text = (
        f"[Artist typo]\n"
        f"Found: {ci['Found Value']}\n"
        f"Suggested: {ci['Suggested Value']}\n"
        f"Similarity: {sim_str}"
    )
    if ci.get("Notes"):
        text += f"\nNotes: {ci['Notes']}"
    text += f"\nCluster {ci['Cluster']}"
    return text


def _format_isrc_issue(ci: dict) -> str:
    text = (
        f"[Duplicate ISRC]\n"
        f"This ISRC appears on multiple rows with different artists.\n"
        f"{ci.get('Notes', '')}\n"
        f"Conflict {ci['Cluster']}"
    )
    return text


def _format_missing_issue(ci: dict) -> str:
    reason = ci.get("Notes", "blank")
    text = (
        f"[Missing required field]\n"
        f"Column: {ci['Column']}\n"
        f"Reason: {reason}\n"
        f"This field is required for ingestion — please fill it in."
    )
    return text


def _format_format_issue(ci: dict) -> str:
    found = str(ci.get("Found Value", ""))
    suggested = str(ci.get("Suggested Value", ""))
    notes = str(ci.get("Notes", ""))
    if ci.get("Excel Row") == 1:
        # Column-wide header annotation.
        text = (
            f"[Format issue — applies to entire column]\n"
            f"Column: {ci['Column']}\n"
            f"{notes}\n"
            f"Suggested fix: {suggested}"
        )
    else:
        text = (
            f"[Format issue]\n"
            f"Column: {ci['Column']}\n"
            f"Found: {found}\n"
            f"{notes}\n"
            f"Suggested: {suggested}"
        )
    return text


def _strip_prior_decoder_annotations(ws) -> int:
    """
    Remove fills + comments previously applied by this scanner so a re-scan can
    re-flag only what's still broken. Touches only:
      • cells whose fill matches one of DECODER_FILL_HEXES, AND
      • comments authored by COMMENT_AUTHOR.
    Anything the user added themselves is left intact.
    Returns the number of cells cleared (for logging).
    """
    cleared = 0
    for row in ws.iter_rows():
        for cell in row:
            stripped = False
            fill = cell.fill
            if fill is not None and fill.fgColor is not None:
                color = (fill.fgColor.value or "").upper()
                # openpyxl prefixes ARGB with "00" sometimes (e.g., "00FCE4D6").
                if any(color.endswith(c) for c in DECODER_FILL_HEXES):
                    cell.fill = PatternFill(fill_type=None)
                    stripped = True
            if cell.comment is not None and cell.comment.author == COMMENT_AUTHOR:
                cell.comment = None
                stripped = True
            if stripped:
                cleared += 1
    return cleared


def _cell_ref(row: int, col: int) -> str:
    """Convert (row=2, col=3) to 'C2' for a threaded-comment cell reference."""
    return f"{get_column_letter(col)}{row}"


def write_annotated_copy(
    input_path: Path, output_path: Path, issues: list[dict], tracks_sheet: str
) -> None:
    """
    Open the original sheet, highlight every flagged cell, and attach an Excel comment
    describing the issue. Saves to output_path.

    If the input already contains a "Decoder Summary" tab (i.e., it's an annotated copy
    being re-scanned), prior decoder fills + comments are cleared first so we don't
    pile new annotations on top of stale ones. The user's own edits are untouched.
    """
    wb = load_workbook(input_path)
    if tracks_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {tracks_sheet!r} not found in {input_path}")
    ws = wb[tracks_sheet]

    is_rescan = DECODER_SUMMARY_SHEET in wb.sheetnames
    if is_rescan:
        _strip_prior_decoder_annotations(ws)
        # Remove the old summary sheet so we can rebuild it cleanly below.
        del wb[DECODER_SUMMARY_SHEET]

    header_to_col: dict[str, int] = {}
    for c_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            header_to_col[str(cell.value)] = c_idx

    # Group issues by cell so a single cell with multiple flagged issues gets one comment.
    cells: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for issue in issues:
        col_idx = header_to_col.get(issue["Column"])
        if col_idx is None:
            continue
        cells[(issue["Excel Row"], col_idx)].append(issue)

    n_artist_cells = 0
    n_isrc_cells = 0
    n_missing_cells = 0
    n_format_cells = 0
    n_mixed_cells = 0
    threaded_targets: list[tuple[str, str]] = []  # (cell_ref, comment_text)

    def _format_one(ci: dict) -> str:
        t = ci.get("Type", TYPE_ARTIST)
        if t == TYPE_ISRC:
            return _format_isrc_issue(ci)
        if t == TYPE_MISSING:
            return _format_missing_issue(ci)
        if t == TYPE_FORMAT:
            return _format_format_issue(ci)
        return _format_artist_issue(ci)

    for (row, col), cell_issues in cells.items():
        cell = ws.cell(row=row, column=col)

        types = {ci.get("Type", TYPE_ARTIST) for ci in cell_issues}
        if types == {TYPE_ARTIST}:
            cell.fill = ARTIST_FILL
            n_artist_cells += 1
        elif types == {TYPE_ISRC}:
            cell.fill = ISRC_FILL
            n_isrc_cells += 1
        elif types == {TYPE_MISSING}:
            cell.fill = MISSING_FILL
            n_missing_cells += 1
        elif types == {TYPE_FORMAT}:
            cell.fill = FORMAT_FILL
            n_format_cells += 1
        else:
            cell.fill = MIXED_FILL
            n_mixed_cells += 1

        if len(cell_issues) == 1:
            text = _format_one(cell_issues[0])
        else:
            parts = ["Multiple issues in this cell:"]
            for ci in cell_issues:
                parts.append("")
                parts.append(_format_one(ci))
            text = "\n".join(parts)

        # Write a legacy comment (Excel "Note") — required as the carrier
        # for the threaded-comment hook. The post-processor below replaces
        # the legacy author with "tc={threadId}" so Excel/Sheets render this
        # as a modern threaded Comment.
        cell.comment = Comment(text, COMMENT_AUTHOR)
        cell.comment.width = 320
        cell.comment.height = 140
        threaded_targets.append((_cell_ref(row, col), text))

    # Decoder Summary sheet at the end -------------------------------------
    summary_ws = wb.create_sheet("Decoder Summary")
    summary_ws["A1"] = "Metadata Decoder — Annotated Copy"
    summary_ws["A1"].font = Font(bold=True, size=14, name="Arial")
    summary_ws["A3"] = "Source"
    summary_ws["B3"] = str(input_path.name)
    summary_ws["A4"] = "Cells flagged (total)"
    summary_ws["B4"] = len(cells)
    summary_ws["A5"] = "  • artist-typo cells"
    summary_ws["B5"] = n_artist_cells
    summary_ws["A6"] = "  • ISRC-conflict cells"
    summary_ws["B6"] = n_isrc_cells
    summary_ws["A7"] = "  • missing-required-field cells"
    summary_ws["B7"] = n_missing_cells
    summary_ws["A8"] = "  • format-issue cells (incl. column headers)"
    summary_ws["B8"] = n_format_cells
    summary_ws["A9"] = "  • cells with multiple issue kinds"
    summary_ws["B9"] = n_mixed_cells
    summary_ws["A10"] = "Total issue records"
    summary_ws["B10"] = len(issues)

    summary_ws["A12"] = (
        "Color key — "
        "light red: artist typo or inconsistency. "
        "Light yellow: ISRC duplicated across rows with different artists. "
        "Light blue: missing required field (blank, whitespace-only, or placeholder like 'N/A'/'TBD'). "
        "Light green: format issue (ISRC/UPC pattern, master split with %, or splits not summing to 100). "
        "When an entire column has a format issue, the green highlight is on the column header (row 1) — "
        "the comment says 'applies to entire column'. "
        "Darker yellow: this cell has more than one kind of problem. "
        "Click any flagged cell to see the threaded comment, or open the Comments side panel "
        "to see them all in a list (Excel: Review › Show Comments / Google Sheets: Comments button at top right). "
        "You can reply to comments and @-tag teammates from there."
    )
    summary_ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")
    summary_ws.column_dimensions["A"].width = 38
    summary_ws.column_dimensions["B"].width = 60
    summary_ws.merge_cells("A12:B17")

    wb.save(output_path)

    # Convert the legacy "Notes" we just wrote into modern threaded
    # "Comments" so Excel / Google Sheets show them in their Comments side
    # panel and let the user reply or @-tag a teammate. openpyxl can't do
    # this natively — we post-process the saved .xlsx zip.
    if threaded_targets:
        try:
            from threaded_comments import inject_threaded_comments
            inject_threaded_comments(
                output_path,
                sheet_name=tracks_sheet,
                author_name=COMMENT_AUTHOR,
                comments=threaded_targets,
            )
        except Exception as e:
            # Threaded conversion is a polish layer — never fatal. The user
            # still has working legacy notes if anything goes wrong.
            print(f"  (threaded-comment conversion skipped: {e})")


# ---- Dashboard --------------------------------------------------------------

import json
from datetime import datetime
from html import escape as _h

DASHBOARD_HISTORY_FILE = "scan_history.json"
DASHBOARD_HTML_FILE = "Decoder Dashboard.html"


def _relpath_from_project(p: Path | None, project_dir: Path) -> str | None:
    """Return p's path relative to project_dir, with forward slashes (browser-friendly).
    For files outside the project folder (e.g., in ~/Downloads), return an
    absolute path so the dashboard's Download link and the Sheets uploader's
    file lookup still resolve correctly."""
    if p is None:
        return None
    try:
        rel = p.resolve().relative_to(project_dir.resolve())
    except ValueError:
        # File lives outside the project folder — store the absolute path so
        # both the dashboard's <a href> and the Sheets uploader can find it.
        return str(p.resolve())
    return rel.as_posix()


def update_dashboard(
    project_dir: Path,
    original_name: str,
    is_rescan: bool,
    stats: dict,
    issues_path: Path | None,
    annotated_path: Path | None,
) -> Path:
    """
    Append/replace a scan entry in scan_history.json and rebuild the HTML dashboard.
    Returns the dashboard path.
    """
    history_path = project_dir / DASHBOARD_HISTORY_FILE
    if history_path.exists():
        try:
            with open(history_path, "r", encoding="utf-8") as f:
                history = json.load(f)
        except Exception:
            history = {}
    else:
        history = {}
    history.setdefault("scans", {})

    entry = {
        "name": original_name,
        "scanned_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "tracks_sheet": stats.get("tracks_sheet", ""),
        "is_rescan": bool(is_rescan),
        "stats": {
            "total_issues": stats.get("total_issues", 0),
            "artist_typo_cells": stats.get("artist_typo_cells", 0),
            "artist_clusters": stats.get("artist_clusters", 0),
            "isrc_conflict_cells": stats.get("isrc_conflict_cells", 0),
            "isrc_conflicts": stats.get("isrc_conflicts", 0),
            "missing_field_issues": stats.get("missing_field_issues", 0),
            "format_issues": stats.get("format_issues", 0),
            "format_columns_wide": stats.get("format_columns_wide", 0),
        },
        # Store paths relative to project_dir so the dashboard's links work
        # even when output files live in subfolders (e.g., "Phase 1/...").
        # Fall back to the bare filename for files outside the project folder.
        "issues_file": _relpath_from_project(issues_path, project_dir),
        "annotated_file": _relpath_from_project(annotated_path, project_dir),
    }
    history["scans"][original_name] = entry
    history["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    with open(history_path, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2)

    dashboard_path = project_dir / DASHBOARD_HTML_FILE
    _write_dashboard_html(dashboard_path, history)
    return dashboard_path


def _pill(label: str, count: int, css_class: str) -> str:
    if count <= 0:
        return ""
    return f'<span class="pill {css_class}">{count} {_h(label)}</span>'


def _row_folder_label(e: dict) -> str:
    """
    Friendly label for the folder a sheet's outputs live in.
    Used for the small subtitle under each dashboard row.
    Falls back to a sensible default for project-root or unknown paths.
    """
    p = e.get("issues_file") or e.get("annotated_file") or ""
    if not p:
        return ""
    parent = str(Path(p).parent)
    if parent in (".", ""):
        return "(project root)"
    # Absolute path → just the last directory component
    name = Path(parent).name
    return name or parent


def _write_dashboard_html(path: Path, history: dict) -> None:
    scans = list(history.get("scans", {}).values())
    # Most recently scanned first.
    scans.sort(key=lambda e: e.get("scanned_at", ""), reverse=True)
    last_updated = history.get("last_updated", "")

    rows_html = []
    for e in scans:
        s = e["stats"]
        total = s["total_issues"]
        if total == 0:
            # Big prominent "All clean!" treatment on zero-issue rows so they
            # stand out as ready-to-ingest at a glance.
            pills = '<span class="all-clean-badge">✓ All clean!</span>'
            total_html = '<span class="total clean-total">—</span>'
        else:
            pills = " ".join(
                p for p in [
                    _pill("artist", s["artist_typo_cells"], "artist"),
                    _pill("ISRC", s["isrc_conflict_cells"], "isrc"),
                    _pill("missing", s["missing_field_issues"], "missing"),
                    _pill("format", s["format_issues"], "format"),
                ] if p
            )
            total_html = f'<span class="total">{total}</span>'

        def _file_row(label: str, filename: str) -> str:
            # Build the Download link's href:
            #   • Relative paths (e.g., "Phase 1/foo.xlsx") stay relative — the
            #     browser resolves them next to the dashboard HTML.
            #   • Absolute paths (e.g., "/Users/me/Downloads/foo.xlsx", used
            #     when the input lives outside the project folder) become a
            #     proper file:// URL with each path segment URL-encoded.
            from urllib.parse import quote as _urlquote
            if filename.startswith("/"):
                href = "file://" + _urlquote(filename)
            else:
                href = _urlquote(filename)
            safe_name = _h(filename)  # used in data-name (sent to local helper)
            safe_href = _h(href)
            return (
                f'<div class="file-row">'
                f'<span class="file-label">{_h(label)}</span>'
                f'<a class="action-btn" href="{safe_href}" download>Download</a>'
                f'<button class="action-btn sheets" type="button" '
                f'data-name="{safe_name}" onclick="openInSheets(this)">Open in Google Sheets</button>'
                f'</div>'
            )

        links = []
        if e.get("issues_file"):
            links.append(_file_row("Issues report", e["issues_file"]))
        if e.get("annotated_file"):
            links.append(_file_row("Annotated copy", e["annotated_file"]))
        links_html = "".join(links) if links else "—"

        rescan_badge = (
            ' <span class="badge">re-scan</span>' if e.get("is_rescan") else ""
        )
        cluster_meta = ""
        if s["artist_clusters"] or s["isrc_conflicts"] or s["format_columns_wide"]:
            parts = []
            if s["artist_clusters"]:
                parts.append(f'{s["artist_clusters"]} artist clusters')
            if s["isrc_conflicts"]:
                parts.append(f'{s["isrc_conflicts"]} ISRC groups')
            if s["format_columns_wide"]:
                parts.append(f'{s["format_columns_wide"]} column-wide format flags')
            cluster_meta = f'<div class="meta">{_h(", ".join(parts))}</div>'

        folder_label = _row_folder_label(e)
        folder_html = (
            f'<div class="folder-subtitle">📁 {_h(folder_label)}</div>'
            if folder_label else ""
        )
        # Searchable text used by the filter bar above the table.
        search_text = " ".join([e.get("name", ""), folder_label,
                                e.get("tracks_sheet") or ""]).lower()
        # The Remove button strips this entry from scan_history.json via the
        # local helper. Files on disk are not touched.
        row_class = " clean-row" if total == 0 else ""
        rows_html.append(f"""
        <tr class="scan-row{row_class}" data-search="{_h(search_text)}" data-name="{_h(e["name"])}">
          <td>
            <div class="sheet-name">{_h(e["name"])}{rescan_badge}</div>
            {folder_html}
            <div class="meta">tab: {_h(e.get("tracks_sheet") or "—")}</div>
            {cluster_meta}
          </td>
          <td><span class="timestamp">{_h(e["scanned_at"])}</span></td>
          <td>{total_html}</td>
          <td>{pills}</td>
          <td class="links">{links_html}</td>
          <td class="remove-cell">
            <button class="remove-btn" type="button" title="Remove this entry from the dashboard (files on disk are not touched)" data-name="{_h(e["name"])}" onclick="removeScan(this)">✕</button>
          </td>
        </tr>
        """)

    if not rows_html:
        body_html = (
            '<div class="empty">No scans yet — run the launcher on a metadata sheet '
            "and this dashboard will populate.</div>"
        )
    else:
        body_html = f"""
        <div class="search-bar">
          <input type="text" id="scan-filter" placeholder="🔍 Filter by sheet name, folder, or tab…" autocomplete="off" />
          <span class="scan-count" id="scan-count">{len(rows_html)} sheet(s)</span>
        </div>
        <table>
          <thead>
            <tr>
              <th>Sheet</th>
              <th>Last scanned</th>
              <th>Total issues</th>
              <th>By check</th>
              <th>Files</th>
              <th></th>
            </tr>
          </thead>
          <tbody id="scan-tbody">
            {"".join(rows_html)}
          </tbody>
        </table>
        """

    script_html = """<script>
// Talks to the local sheets_uploader.py helper (started in the background by
// "Scan Metadata Sheet.command"). The helper uploads the .xlsx to Drive,
// converts it to a Google Sheet, and returns the URL.
var UPLOADER_URL = 'http://127.0.0.1:53127';

function openInSheets(btn) {
  var name = btn.getAttribute('data-name') || 'the file';
  if (btn.dataset.busy === '1') return;
  btn.dataset.busy = '1';
  var originalText = btn.textContent;
  btn.textContent = 'Uploading…';
  btn.style.opacity = '0.7';

  var done = function() {
    btn.dataset.busy = '';
    btn.textContent = originalText;
    btn.style.opacity = '';
  };

  showToast('Uploading <strong>' + escapeHtml(name) +
            '</strong> to Google Sheets…');

  fetch(UPLOADER_URL + '/upload?file=' + encodeURIComponent(name), {
    method: 'POST'
  }).then(function(r) {
    return r.json().then(function(data) { return { ok: r.ok, data: data }; });
  }).then(function(res) {
    if (res.ok && res.data.url) {
      window.open(res.data.url, '_blank');
      showToast('Opened <strong>' + escapeHtml(name) + '</strong> in Google Sheets.');
    } else {
      var msg = (res.data && res.data.error) ? res.data.error : 'Upload failed.';
      var html = 'Couldn\\'t upload: ' + escapeHtml(msg);
      // Common case: missing credentials.json — point to the setup guide.
      if (/credentials\\.json/i.test(msg)) {
        html += '<br>See <em>Google Sheets Upload Setup.md</em> in the project folder.';
      }
      showToast(html, 10000);
    }
  }).catch(function() {
    // Network error → helper isn't running.
    showToast('Upload helper not running. Re-run <em>Scan Metadata Sheet.command</em>' +
              ' to start it, then try again.<br>' +
              '(If that doesn\\'t help, open this dashboard in Chrome or Firefox.)',
              12000);
  }).then(done, done);
}

function escapeHtml(s) {
  return String(s).replace(/[&<>"']/g, function(c) {
    return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];
  });
}
function showToast(html, ms) {
  var t = document.getElementById('cw-toast');
  if (!t) {
    t = document.createElement('div');
    t.id = 'cw-toast';
    t.className = 'toast';
    document.body.appendChild(t);
  }
  t.innerHTML = html;
  void t.offsetWidth;
  t.classList.add('show');
  clearTimeout(t._hideTimer);
  t._hideTimer = setTimeout(function() { t.classList.remove('show'); },
                            ms || 7000);
}

// ---- Search/filter bar -------------------------------------------------
// Filters visible rows in real-time by the data-search attribute on each
// <tr>. Persists the query in sessionStorage so it survives page reloads.
function applyScanFilter(q) {
  var rows = document.querySelectorAll('tr.scan-row');
  var visible = 0;
  q = (q || '').trim().toLowerCase();
  rows.forEach(function(r) {
    var hay = r.getAttribute('data-search') || '';
    var match = !q || hay.indexOf(q) !== -1;
    r.classList.toggle('hidden', !match);
    if (match) visible++;
  });
  var counter = document.getElementById('scan-count');
  if (counter) {
    counter.textContent = q
      ? visible + ' of ' + rows.length + ' sheet(s)'
      : rows.length + ' sheet(s)';
  }
}
(function initFilter() {
  var input = document.getElementById('scan-filter');
  if (!input) return;
  var saved = '';
  try { saved = sessionStorage.getItem('decoder.filter') || ''; } catch(e) {}
  if (saved) {
    input.value = saved;
    applyScanFilter(saved);
  }
  input.addEventListener('input', function() {
    var v = input.value;
    try { sessionStorage.setItem('decoder.filter', v); } catch(e) {}
    applyScanFilter(v);
  });
})();

// ---- Remove-from-dashboard button --------------------------------------
// Calls the local helper to delete the entry from scan_history.json.
// Files on disk are NOT touched.
function removeScan(btn) {
  var name = btn.getAttribute('data-name');
  if (!name) return;
  if (btn.dataset.busy === '1') return;
  if (!confirm('Remove "' + name + '" from the dashboard?\\n\\n' +
               'This only removes the dashboard entry. ' +
               'The .xlsx files on your Mac stay where they are.')) {
    return;
  }
  btn.dataset.busy = '1';
  btn.style.opacity = '0.5';

  fetch(UPLOADER_URL + '/delete-scan?name=' + encodeURIComponent(name), {
    method: 'POST'
  }).then(function(r) {
    return r.json().then(function(data) { return { ok: r.ok, data: data }; });
  }).then(function(res) {
    if (res.ok) {
      // Drop the row from the visible table immediately.
      var row = btn.closest('tr');
      if (row) row.parentNode.removeChild(row);
      // Re-run the filter so the count and visible-state update.
      var input = document.getElementById('scan-filter');
      applyScanFilter(input ? input.value : '');
      showToast('Removed <strong>' + escapeHtml(name) + '</strong> from the dashboard.');
    } else {
      var msg = (res.data && res.data.error) ? res.data.error : 'Remove failed.';
      showToast("Couldn't remove: " + escapeHtml(msg), 8000);
      btn.dataset.busy = '';
      btn.style.opacity = '';
    }
  }).catch(function() {
    showToast('Remove helper not running. Re-run <em>Scan Metadata Sheet.command</em>' +
              ' to start it, then try again.', 10000);
    btn.dataset.busy = '';
    btn.style.opacity = '';
  });
}
</script>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="cache-control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="pragma" content="no-cache">
<meta http-equiv="expires" content="0">
<title>Metadata Decoder Dashboard</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", system-ui, sans-serif;
         max-width: 1200px; margin: 2em auto; padding: 0 2em; color: #2c3e50;
         background: #FAFBFC; }}
  h1 {{ font-size: 1.9em; margin: 0 0 0.2em; color: #1F4E78; }}
  .subtitle {{ color: #6c757d; font-size: 0.95em; margin: 0 0 1.6em; }}
  table {{ width: 100%; border-collapse: collapse; background: #fff;
           box-shadow: 0 1px 3px rgba(0,0,0,0.05); border-radius: 6px; overflow: hidden; }}
  th {{ text-align: left; padding: 0.7em 1em; background: #1F4E78; color: #FFF;
        font-weight: 600; font-size: 0.85em; letter-spacing: 0.02em; text-transform: uppercase; }}
  td {{ padding: 0.85em 1em; border-bottom: 1px solid #f0f0f0; font-size: 0.93em;
        vertical-align: top; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: #F5F8FB; }}
  .total {{ font-weight: 700; font-size: 1.1em; color: #2c3e50; }}
  .pill {{ display: inline-block; padding: 0.2em 0.7em; border-radius: 999px;
          font-size: 0.78em; margin-right: 0.35em; margin-bottom: 0.2em; font-weight: 500; }}
  .artist {{ background: #FCE4D6; color: #842029; }}
  .isrc {{ background: #FFF2CC; color: #664d03; }}
  .missing {{ background: #DDEBF7; color: #084298; }}
  .format {{ background: #C6EFCE; color: #0F5132; }}
  .clean {{ background: #D1E7DD; color: #0F5132; font-weight: 600; }}
  /* Big "All clean!" treatment for zero-issue rows. */
  .all-clean-badge {{ display: inline-block; padding: 0.4em 0.9em; background: #D1E7DD;
                      color: #0F5132; border-radius: 999px; font-size: 0.95em;
                      font-weight: 700; letter-spacing: 0.01em;
                      box-shadow: 0 0 0 2px #b8dec6 inset; }}
  .clean-row td {{ background: #FAFFF8; }}
  .clean-row:hover td {{ background: #F0F9EE; }}
  .clean-total {{ color: #b8c5b5; font-weight: 500; font-size: 1em; }}
  /* Folder subtitle under each sheet name. */
  .folder-subtitle {{ color: #6c757d; font-size: 0.82em; margin-top: 0.2em;
                      font-weight: 500; }}
  /* Search bar above the table. */
  .search-bar {{ display: flex; align-items: center; gap: 1em; margin-bottom: 0.8em; }}
  .search-bar input {{ flex: 1; padding: 0.55em 0.9em; font-size: 0.95em;
                       border: 1px solid #d0d7de; border-radius: 6px;
                       font-family: inherit; background: #fff;
                       box-shadow: 0 1px 2px rgba(0,0,0,0.03); }}
  .search-bar input:focus {{ outline: none; border-color: #1F4E78;
                              box-shadow: 0 0 0 3px rgba(31,78,120,0.1); }}
  .scan-count {{ color: #6c757d; font-size: 0.85em; white-space: nowrap; }}
  /* Remove (✕) button on each row. */
  .remove-cell {{ padding-left: 0.3em; padding-right: 0.6em; text-align: right; }}
  .remove-btn {{ display: inline-flex; align-items: center; justify-content: center;
                 width: 24px; height: 24px; padding: 0;
                 background: transparent; border: 1px solid transparent;
                 border-radius: 4px; color: #adb5bd; cursor: pointer;
                 font-size: 0.85em; line-height: 1; font-family: inherit;
                 transition: all 0.15s; }}
  .remove-btn:hover {{ background: #FCE4D6; border-color: #f1aeb5;
                       color: #842029; }}
  .scan-row.hidden {{ display: none; }}
  .links a {{ color: #0d6efd; text-decoration: none; font-weight: 500; }}
  .links a:hover {{ text-decoration: underline; }}
  .file-row {{ display: flex; align-items: center; gap: 0.4em;
               margin-bottom: 0.35em; flex-wrap: wrap; }}
  .file-row:last-child {{ margin-bottom: 0; }}
  .file-label {{ color: #2c3e50; font-weight: 500; min-width: 7em; font-size: 0.9em; }}
  .action-btn {{ display: inline-flex; align-items: center; gap: 0.3em;
                 padding: 0.25em 0.65em; background: #fff; border: 1px solid #d0d7de;
                 border-radius: 4px; color: #0d6efd; text-decoration: none;
                 font-size: 0.82em; cursor: pointer; font-family: inherit; line-height: 1.4; }}
  .action-btn:hover {{ background: #F5F8FB; border-color: #0d6efd; text-decoration: none; }}
  .action-btn.sheets {{ color: #0F5132; }}
  .action-btn.sheets:hover {{ border-color: #0F5132; background: #F1F8F4; }}
  .toast {{ position: fixed; bottom: 1.6em; right: 1.6em; background: #1F4E78; color: #fff;
            padding: 0.9em 1.2em; border-radius: 6px; box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            font-size: 0.88em; max-width: 320px; opacity: 0; transform: translateY(8px);
            transition: opacity 0.25s, transform 0.25s; pointer-events: none;
            z-index: 1000; line-height: 1.5; }}
  .toast.show {{ opacity: 1; transform: translateY(0); }}
  .timestamp {{ color: #6c757d; font-size: 0.9em; white-space: nowrap; }}
  .sheet-name {{ font-weight: 600; color: #2c3e50; }}
  .meta {{ color: #6c757d; font-size: 0.82em; margin-top: 0.15em; }}
  .badge {{ display: inline-block; padding: 0.1em 0.5em; background: #E7F1FF;
            color: #084298; border-radius: 4px; font-size: 0.75em; font-weight: 500;
            margin-left: 0.3em; vertical-align: middle; }}
  .empty {{ padding: 3em; text-align: center; color: #6c757d; background: #fff;
            border-radius: 6px; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }}
  footer {{ color: #95a5a6; font-size: 0.83em; margin-top: 2.5em;
            padding-top: 1em; border-top: 1px solid #E5E8EB; line-height: 1.6; }}
</style>
</head>
<body>
  <h1>Metadata Decoder Dashboard</h1>
  <p class="subtitle">Last refreshed: {_h(last_updated)}</p>
  {body_html}
  <footer>
    For each report, choose <strong>Download</strong> to save the .xlsx locally, or <strong>Open in Google Sheets</strong> to upload it to your Drive.<br>
    Run <em>Scan Metadata Sheet.command</em> on a sheet to refresh this dashboard. Re-scanning an annotated copy updates that file in place — color key:
    <span class="pill artist">artist typo</span>
    <span class="pill isrc">ISRC conflict</span>
    <span class="pill missing">missing field</span>
    <span class="pill format">format issue</span>
  </footer>
  {script_html}
</body>
</html>
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


# ---- Main -------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description="Scan a metadata sheet for artist-name issues.")
    parser.add_argument("input", help="Path to the metadata .xlsx file.")
    parser.add_argument(
        "--mode",
        choices=["report", "annotated", "both"],
        default="report",
        help=(
            "report = separate issues spreadsheet (default); "
            "annotated = marked-up copy of the original; "
            "both = produce both files."
        ),
    )
    parser.add_argument(
        "--output",
        help=(
            "Optional output path. With --mode both, this is treated as a directory. "
            "Defaults: <input>_issues.xlsx and/or <input>_annotated.xlsx alongside the input."
        ),
    )
    args = parser.parse_args()

    in_path = Path(args.input).expanduser().resolve()
    if not in_path.exists():
        print(f"Input file not found: {in_path}", file=sys.stderr)
        return 1

    # Detect re-scan: when the input is itself an annotated copy from a previous run.
    # In that case the issues report keeps its name based on the *original* sheet, and
    # the annotated copy is updated in place. This way the user can iterate on a
    # single file across multiple scans.
    stem = in_path.stem
    is_rescan = stem.endswith("_annotated")
    base_stem = stem[: -len("_annotated")] if is_rescan else stem

    # Predict where the issues file will land so analyze() can carry forward
    # any user-typed Corrections from a previous run.
    prior_issues_path = in_path.with_name(base_stem + "_issues.xlsx")

    (
        issues,
        cluster_summary,
        isrc_summary,
        missing_summary,
        missing_per_cell,
        format_column_summary,
        format_row_summary,
        format_corrections,
        splits_review,
        stats,
        split_errors,
        id_mismatches,
    ) = analyze(in_path, issues_output_path=prior_issues_path)

    outputs_written: list[Path] = []

    if args.mode in ("report", "both"):
        if args.mode == "report" and args.output:
            report_path = Path(args.output).expanduser().resolve()
        else:
            report_path = in_path.with_name(base_stem + "_issues.xlsx")
        write_report(
            report_path,
            issues,
            cluster_summary,
            isrc_summary,
            missing_summary,
            missing_per_cell,
            format_column_summary,
            format_row_summary,
            format_corrections,
            splits_review,
            in_path,
            split_errors=split_errors,
            id_mismatches=id_mismatches,
            detected_format=stats.get("detected_format", ""),
        )
        outputs_written.append(report_path)

    annotated_path: Path | None = None
    if args.mode in ("annotated", "both"):
        if args.mode == "annotated" and args.output:
            annotated_path = Path(args.output).expanduser().resolve()
        elif is_rescan:
            # Re-scan: write back to the same file the user picked.
            annotated_path = in_path
        else:
            annotated_path = in_path.with_name(base_stem + "_annotated.xlsx")
        write_annotated_copy(in_path, annotated_path, issues, stats["tracks_sheet"])
        outputs_written.append(annotated_path)

    # The dashboard always lives next to scan_metadata.py (the project root),
    # not next to the input file — so scanning a sheet inside a sub-folder
    # like "Phase 1/" still updates the project-level dashboard.
    project_root = Path(__file__).resolve().parent
    update_dashboard(
        project_dir=project_root,
        original_name=base_stem,
        is_rescan=is_rescan,
        stats=stats,
        issues_path=outputs_written[0] if args.mode in ("report", "both") else None,
        annotated_path=annotated_path,
    )

    print("Scan complete.")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    for p in outputs_written:
        print(f"  wrote: {p}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
