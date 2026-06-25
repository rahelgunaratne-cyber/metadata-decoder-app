"""
Apply ISRC-conflict review entries from an Issues file to its matching
Annotated copy, then log every change.

Workflow:
    1. The user fills in two columns on the ISRC Conflicts tab of
       <name>_issues.xlsx (one row per offending row, grouped by Conflict ID):
         • "Confirm OK?"     — type any of {OK, yes, y, x, ✓, true, 1} to
                                mark this duplicate as intentional. Future
                                scans will skip this exact ISRC + artist
                                constellation (persisted to .isrc_leave.json
                                at the project root).
         • "Corrected ISRC"  — type the replacement ISRC for that one row.
                                The annotated workbook's Track ISRC cell on
                                that row gets overwritten in place.
       Empty in both columns → defer the decision; nothing happens.
       Both filled in → "Corrected ISRC" wins (the row gets corrected; OK is
       ignored for that row, with a warning in the audit log).
    2. Run this script with <name>_annotated.xlsx as the argument
       (typically via "Apply ISRC Corrections.command").
    3. The annotated copy is rewritten in place with the corrections applied.
    4. An "Applied ISRC Corrections" tab is appended/refreshed in the issues
       file with a row per cell that changed (sheet, row, before, after) and
       a section listing the conflicts marked OK this run.
    5. For any conflict where 2+ rows were confirmed OK and still share the
       same ISRC after corrections, a leave record is added to
       .isrc_leave.json so the scanner stops flagging that constellation.

Usage:
    python3 apply_isrc_corrections.py <annotated.xlsx>
    python3 apply_isrc_corrections.py <annotated.xlsx> --dry-run
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Reuse scanner helpers and constants so behavior matches the scan exactly.
from scan_metadata import (
    ISRC_COLUMN,
    ISRC_LEAVE_RECORDS_FILENAME,
    ISRC_PATTERN,
    detect_tracks_sheet,
    is_confirm_ok,
    load_isrc_leave_records,
    normalize,
    save_isrc_leave_records,
)


# ---- Reading the ISRC Conflicts tab ----------------------------------------


def read_review_rows(issues_path: Path) -> list[dict]:
    """
    Parse the ISRC Conflicts tab into a list of per-row review records:
        {
            "conflict_id":     "I3",
            "isrc":            "ESA021601309",
            "excel_row":       42,
            "title":           "...",
            "artist":          "Eloi Vicente",
            "confirm_ok_raw":  ""  (or "OK"/"yes"/etc.),
            "confirm_ok":      bool,
            "corrected_isrc":  ""  (or e.g. "ESA021601400"),
        }
    """
    wb = load_workbook(issues_path, read_only=True, data_only=True)
    if "ISRC Conflicts" not in wb.sheetnames:
        raise SystemExit(
            f"No 'ISRC Conflicts' tab in {issues_path.name}. "
            "Run a scan first so the corrections worksheet is generated."
        )
    ws = wb["ISRC Conflicts"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    needed = {
        "Conflict",
        "ISRC",
        "Excel Row",
        "Track Title",
        "Track Display Artist",
        "Confirm OK?",
        "Corrected ISRC",
    }
    missing = needed - set(header)
    if missing:
        raise SystemExit(
            f"'ISRC Conflicts' tab is missing column(s): {sorted(missing)}. "
            "Re-run the scan to regenerate the issues file with the new layout."
        )
    idx = {name: header.index(name) for name in needed}
    out: list[dict] = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        try:
            excel_row = int(r[idx["Excel Row"]])
        except (TypeError, ValueError):
            continue
        ok_raw = "" if r[idx["Confirm OK?"]] is None else str(r[idx["Confirm OK?"]]).strip()
        corr_raw = "" if r[idx["Corrected ISRC"]] is None else str(r[idx["Corrected ISRC"]]).strip()
        out.append({
            "conflict_id":    str(r[idx["Conflict"]] or "").strip(),
            "isrc":           str(r[idx["ISRC"]] or "").strip(),
            "excel_row":      excel_row,
            "title":          str(r[idx["Track Title"]] or "").strip(),
            "artist":         str(r[idx["Track Display Artist"]] or "").strip(),
            "confirm_ok_raw": ok_raw,
            "confirm_ok":     is_confirm_ok(ok_raw),
            "corrected_isrc": corr_raw,
        })
    return out


# ---- Planning corrections + leave records ----------------------------------


def _validate_isrc(value: str) -> bool:
    """Reuse the scanner's format check (12-char canonical form, no hyphens)."""
    cleaned = value.replace("-", "").upper().strip()
    return bool(ISRC_PATTERN.match(cleaned))


def plan_actions(review: list[dict]) -> tuple[list[dict], list[dict], list[str]]:
    """
    Returns:
        replacements   — list of per-row corrections to apply, each:
                            {excel_row, old_isrc, new_isrc, conflict_id,
                             artist, title, format_warning}
        leave_records  — list of conflict-level leave records to persist:
                            {conflict_id, isrc, signature, rows, note, added_at}
        warnings       — list of human-readable warning strings.
    """
    warnings: list[str] = []
    replacements: list[dict] = []

    # Group rows by conflict so we can decide leave records per-conflict.
    by_conflict: dict[str, list[dict]] = defaultdict(list)
    for r in review:
        by_conflict[r["conflict_id"]].append(r)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    leave_records: list[dict] = []

    for conflict_id, rows_in in by_conflict.items():
        for r in rows_in:
            has_corr = bool(r["corrected_isrc"])
            has_ok = r["confirm_ok"]

            if has_corr and has_ok:
                warnings.append(
                    f"{conflict_id} row {r['excel_row']}: both 'Confirm OK?' "
                    f"({r['confirm_ok_raw']!r}) and 'Corrected ISRC' "
                    f"({r['corrected_isrc']!r}) were filled in. Applying the "
                    "correction; the OK marker is ignored for this row."
                )

            if has_corr:
                new_isrc = r["corrected_isrc"]
                fmt_ok = _validate_isrc(new_isrc)
                if not fmt_ok:
                    warnings.append(
                        f"{conflict_id} row {r['excel_row']}: Corrected ISRC "
                        f"{new_isrc!r} doesn't match the standard 12-character "
                        "format (CCXXXYYNNNNN). Applied anyway — please verify."
                    )
                replacements.append({
                    "excel_row":      r["excel_row"],
                    "old_isrc":       r["isrc"],
                    "new_isrc":       new_isrc,
                    "conflict_id":    conflict_id,
                    "artist":         r["artist"],
                    "title":          r["title"],
                    "format_warning": not fmt_ok,
                })

        # Determine which rows in this conflict end up suppressed:
        # rows marked OK that AREN'T being corrected. If 2+ such rows remain
        # sharing the same ISRC with different artists, persist a leave record
        # for that residual sub-conflict.
        ok_rows = [
            r for r in rows_in
            if r["confirm_ok"] and not r["corrected_isrc"]
        ]
        # Group OK rows by their (still-current) ISRC.
        by_isrc: dict[str, list[dict]] = defaultdict(list)
        for r in ok_rows:
            by_isrc[r["isrc"]].append(r)

        for isrc, group in by_isrc.items():
            # We need at least 2 rows AND 2+ distinct artists for there to be
            # a real residual conflict worth suppressing.
            distinct_artists = {normalize(g["artist"]) for g in group if g["artist"]}
            if len(group) < 2 or len(distinct_artists) < 2:
                continue
            signature = sorted({
                (isrc.strip().upper(), normalize(g["artist"]))
                for g in group if g["artist"]
            })
            leave_records.append({
                "conflict_id": conflict_id,
                "isrc":        isrc,
                "signature":   signature,
                "rows":        sorted(g["excel_row"] for g in group),
                "note":        "Confirmed via 'Confirm OK?' column.",
                "added_at":    stamp,
            })

    return replacements, leave_records, warnings


# ---- Applying replacements to the annotated workbook ----------------------


def apply_to_workbook(annotated_path: Path, replacements: list[dict]) -> list[dict]:
    """
    Open the annotated workbook, find the tracks sheet, and overwrite the
    Track ISRC cell on each replacement's row. Returns an audit log.
    """
    if not replacements:
        return []

    tracks_sheet, _ = detect_tracks_sheet(annotated_path)
    wb = load_workbook(annotated_path)
    if tracks_sheet not in wb.sheetnames:
        raise SystemExit(f"Tracks sheet {tracks_sheet!r} not found in workbook.")
    ws = wb[tracks_sheet]

    # Locate the Track ISRC column from the header row.
    header_row = [
        (c.value if c.value is not None else "")
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    try:
        isrc_col_idx = next(
            j for j, h in enumerate(header_row, start=1)
            if isinstance(h, str) and h == ISRC_COLUMN
        )
    except StopIteration:
        raise SystemExit(
            f"Couldn't find the {ISRC_COLUMN!r} column in {tracks_sheet!r}."
        )

    audit: list[dict] = []
    for r in replacements:
        cell = ws.cell(row=r["excel_row"], column=isrc_col_idx)
        before = "" if cell.value is None else str(cell.value)
        cell.value = r["new_isrc"]
        audit.append({
            "Sheet":       tracks_sheet,
            "Excel Row":   r["excel_row"],
            "Conflict":    r["conflict_id"],
            "Artist":      r["artist"],
            "Title":       r["title"],
            "Before":      before,
            "After":       r["new_isrc"],
            "Format note": "non-standard format" if r["format_warning"] else "",
        })

    if audit:
        wb.save(annotated_path)
    return audit


# ---- Audit-log tab in the issues file --------------------------------------


HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")


def write_audit_log(
    issues_path: Path,
    audit: list[dict],
    leave_records: list[dict],
    warnings: list[str],
    annotated_path: Path,
) -> None:
    """Append/replace the 'Applied ISRC Corrections' tab on the issues file."""
    wb = load_workbook(issues_path)
    name = "Applied ISRC Corrections"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([f"Applied {len(audit)} ISRC change(s) at {stamp}"])
    ws["A1"].font = Font(bold=True, size=12, name="Arial")
    ws.append([f"Annotated copy: {annotated_path.name}"])
    ws.append([f"Conflicts confirmed OK this run: {len(leave_records)}"])
    ws.append([f"Warnings: {len(warnings)}"])
    ws.append([])

    headers = ["Sheet", "Excel Row", "Conflict", "Artist", "Title", "Before", "After", "Format note"]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    for i, ch in enumerate(audit, start=1):
        ws.append([ch[h] for h in headers])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = ALT_FILL

    widths = {1: 22, 2: 10, 3: 10, 4: 25, 5: 30, 6: 18, 7: 18, 8: 22}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    if leave_records:
        ws.append([])
        ws.append(["Conflicts confirmed OK this run (persisted to .isrc_leave.json):"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, name="Arial")
        ws.append(["Conflict", "ISRC", "Rows", "Artists"])
        hdr = ws.max_row
        for cell in ws[hdr]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for r in leave_records:
            artists = "; ".join(sorted({a for _, a in r["signature"]}))
            ws.append([
                r.get("conflict_id", ""),
                r.get("isrc", ""),
                ", ".join(str(x) for x in r.get("rows", [])),
                artists,
            ])

    if warnings:
        ws.append([])
        ws.append(["Warnings:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, name="Arial")
        for w_str in warnings:
            ws.append([w_str])

    wb.save(issues_path)


# ---- Persistence helpers ---------------------------------------------------


def merge_leave_records(project_dir: Path, new_records: list[dict]) -> int:
    """
    Add new ISRC LEAVE entries to .isrc_leave.json. De-dupes by signature
    (set of (isrc, normalized_artist) pairs). Returns the count of newly
    added records.
    """
    existing_sigs = set(load_isrc_leave_records(project_dir))  # frozensets of (isrc, norm)
    path = project_dir / ISRC_LEAVE_RECORDS_FILENAME
    existing_records: list[dict] = []
    if path.exists():
        try:
            existing_records = json.loads(path.read_text()).get("records", [])
        except Exception:
            existing_records = []

    added = 0
    for rec in new_records:
        sig = frozenset(tuple(p) for p in rec.get("signature", []))
        if not sig or sig in existing_sigs:
            continue
        existing_records.append({
            "isrc":      rec.get("isrc", ""),
            "signature": [list(p) for p in rec.get("signature", [])],
            "rows":      list(rec.get("rows", [])),
            "note":      rec.get("note", ""),
            "added_at":  rec.get("added_at", ""),
        })
        existing_sigs.add(sig)
        added += 1

    save_isrc_leave_records(project_dir, existing_records)
    return added


# ---- Main ------------------------------------------------------------------


def derive_issues_path(annotated_path: Path) -> Path:
    """<name>_annotated.xlsx → <name>_issues.xlsx (sibling)."""
    stem = annotated_path.stem
    if not stem.endswith("_annotated"):
        raise SystemExit(
            "Expected an annotated file (filename should end in '_annotated.xlsx'). "
            f"Got: {annotated_path.name}"
        )
    base = stem[: -len("_annotated")]
    return annotated_path.with_name(base + "_issues.xlsx")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Apply 'Confirm OK?' / 'Corrected ISRC' entries from "
            "<name>_issues.xlsx to <name>_annotated.xlsx."
        )
    )
    parser.add_argument("annotated", help="Path to the <name>_annotated.xlsx file.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Compute changes but don't write any files.")
    args = parser.parse_args()

    annotated_path = Path(args.annotated).expanduser().resolve()
    if not annotated_path.exists():
        print(f"File not found: {annotated_path}", file=sys.stderr)
        return 1

    issues_path = derive_issues_path(annotated_path)
    if not issues_path.exists():
        print(
            f"Issues file not found: {issues_path}\n"
            "Run a scan first so the corrections worksheet is generated.",
            file=sys.stderr,
        )
        return 1

    review = read_review_rows(issues_path)
    if not review:
        print("No ISRC conflict rows found in the issues file — nothing to do.")
        return 0

    replacements, leave_records, warnings = plan_actions(review)

    print(f"ISRC conflict rows total: {len(review)}")
    print(f"  → corrections to apply:        {len(replacements)}")
    print(f"  → conflicts confirmed OK:      {len(leave_records)}")
    print(f"  → warnings:                    {len(warnings)}")

    if warnings:
        print("\nWarnings:")
        for w in warnings:
            print(f"  ! {w}")

    if args.dry_run:
        print("\n[dry-run] would change Track ISRC on these rows:")
        for r in replacements:
            print(
                f"   row {r['excel_row']:>4}  {r['conflict_id']:<5}  "
                f"{r['old_isrc']!r} → {r['new_isrc']!r}  ({r['artist']})"
            )
        if leave_records:
            print("\n[dry-run] would add these LEAVE records:")
            for r in leave_records:
                print(
                    f"   {r['conflict_id']}  ISRC {r['isrc']}  "
                    f"rows {r['rows']}"
                )
        return 0

    audit = apply_to_workbook(annotated_path, replacements)
    print(f"\nCell changes written to {annotated_path.name}: {len(audit)}")

    write_audit_log(issues_path, audit, leave_records, warnings, annotated_path)
    print(f"Audit log written to {issues_path.name} → 'Applied ISRC Corrections' tab.")

    if leave_records:
        # Project-level so a single decision applies across phase folders,
        # mirroring how .artist_leave.json works.
        project_root = Path(__file__).resolve().parent
        added = merge_leave_records(project_root, leave_records)
        print(f"LEAVE markers stored in .isrc_leave.json (newly added: {added}).")

    return 0


if __name__ == "__main__":
    sys.exit(main())
