"""
Apply Missing Field Corrections from an Issues file to its matching
Annotated copy, then log every change.

Workflow:
    1. The user fills in the "Fill Value" column on the Missing Field
       Corrections tab of <name>_issues.xlsx (one row per missing cell).
       The column is pre-filled with "Suggested Fill" where the scanner
       can guess confidently — release-level fields with consistent
       siblings, sheet-level fields like Track Language with one prevailing
       value, Track Display Artist when Artist 1 Name on Track is filled.
       - Empty Fill Value → skip this cell (defer the decision).
       - A value          → write that value to the (Excel Row, Column) cell
                            in the annotated copy.
    2. Run this script with <name>_annotated.xlsx as the argument
       (typically via "Apply Missing Field Corrections.command").
    3. The annotated copy is rewritten in place with the fills applied.
    4. An "Applied Missing Field Corrections" tab is appended/refreshed in
       the issues file with a row per cell that changed (sheet, row,
       column, before, after).

Usage:
    python3 apply_missing_corrections.py <annotated.xlsx>
    python3 apply_missing_corrections.py <annotated.xlsx> --dry-run
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scan_metadata import detect_tracks_sheet


# ---- Reading the Missing Field Corrections tab ----------------------------


def read_fill_rows(issues_path: Path) -> list[dict]:
    """
    Parse the Missing Field Corrections tab into a list of fill records:
        {
            "excel_row":  141,
            "column":     "Track P Line",
            "title":      "Will You Be There",
            "artist":     "Sáloa Farah",
            "reason":     "blank",
            "suggested":  "Bossa Nova 58",
            "fill_value": "Bossa Nova 58",
            "source":     "consistent across other rows with UPC ..."
        }
    """
    wb = load_workbook(issues_path, read_only=True, data_only=True)
    if "Missing Field Corrections" not in wb.sheetnames:
        raise SystemExit(
            f"No 'Missing Field Corrections' tab in {issues_path.name}. "
            "Run a scan first so the corrections worksheet is generated."
        )
    ws = wb["Missing Field Corrections"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    needed = {
        "Excel Row",
        "Column",
        "Track Title",
        "Track Display Artist",
        "Reason for missing",
        "Suggested Fill",
        "Fill Value",
        "Suggestion source",
    }
    missing = needed - set(header)
    if missing:
        raise SystemExit(
            f"'Missing Field Corrections' tab is missing column(s): "
            f"{sorted(missing)}. Re-run the scan to regenerate the issues "
            "file with the new layout."
        )
    idx = {name: header.index(name) for name in needed}
    out: list[dict] = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        try:
            excel_row = int(r[idx["Excel Row"]])
        except (TypeError, ValueError):
            continue
        colname = str(r[idx["Column"]] or "").strip()
        if not colname:
            continue
        out.append({
            "excel_row":  excel_row,
            "column":     colname,
            "title":      str(r[idx["Track Title"]] or "").strip(),
            "artist":     str(r[idx["Track Display Artist"]] or "").strip(),
            "reason":     str(r[idx["Reason for missing"]] or "").strip(),
            "suggested":  "" if r[idx["Suggested Fill"]] is None else str(r[idx["Suggested Fill"]]).strip(),
            "fill_value": "" if r[idx["Fill Value"]] is None else str(r[idx["Fill Value"]]).strip(),
            "source":     str(r[idx["Suggestion source"]] or "").strip(),
        })
    return out


# ---- Applying fills to the annotated workbook ----------------------------


def apply_to_workbook(annotated_path: Path, fills: list[dict]) -> list[dict]:
    """
    Open the annotated workbook, find the tracks sheet, and write each fill
    value into the (Excel Row, Column) cell. Returns an audit log of every
    cell that actually changed.
    """
    if not fills:
        return []

    tracks_sheet, _ = detect_tracks_sheet(annotated_path)
    wb = load_workbook(annotated_path)
    if tracks_sheet not in wb.sheetnames:
        raise SystemExit(f"Tracks sheet {tracks_sheet!r} not found in workbook.")
    ws = wb[tracks_sheet]

    # Map column header → 1-based column index.
    header_row = [
        (c.value if c.value is not None else "")
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    col_to_idx: dict[str, int] = {}
    for j, h in enumerate(header_row, start=1):
        if isinstance(h, str):
            col_to_idx[h] = j

    audit: list[dict] = []
    skipped_unknown_cols: list[str] = []
    for f in fills:
        col_idx = col_to_idx.get(f["column"])
        if col_idx is None:
            skipped_unknown_cols.append(f["column"])
            continue
        cell = ws.cell(row=f["excel_row"], column=col_idx)
        before = "" if cell.value is None else str(cell.value)
        cell.value = f["fill_value"]
        audit.append({
            "Sheet":       tracks_sheet,
            "Excel Row":   f["excel_row"],
            "Column":      f["column"],
            "Track Title": f["title"],
            "Artist":      f["artist"],
            "Before":      before,
            "After":       f["fill_value"],
            "Source":      f["source"] or ("user-typed" if f["fill_value"] != f["suggested"] else "auto-suggested"),
        })

    if audit:
        wb.save(annotated_path)
    return audit, skipped_unknown_cols


# ---- Audit-log tab in the issues file ------------------------------------


HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")


def write_audit_log(
    issues_path: Path,
    audit: list[dict],
    skipped_unknown_cols: list[str],
    annotated_path: Path,
) -> None:
    """Append/replace the 'Applied Missing Fields' tab on the issues file."""
    wb = load_workbook(issues_path)
    # Keep tab name ≤ 31 chars (Excel's limit).
    name = "Applied Missing Fields"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([f"Applied {len(audit)} cell fill(s) at {stamp}"])
    ws["A1"].font = Font(bold=True, size=12, name="Arial")
    ws.append([f"Annotated copy: {annotated_path.name}"])
    if skipped_unknown_cols:
        ws.append([
            f"Skipped {len(skipped_unknown_cols)} entry(ies) — "
            f"column(s) not found in tracks sheet: {sorted(set(skipped_unknown_cols))}"
        ])
    ws.append([])

    headers = ["Sheet", "Excel Row", "Column", "Track Title", "Artist", "Before", "After", "Source"]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    for i, ch in enumerate(audit, start=1):
        ws.append([ch[h] for h in headers])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = ALT_FILL

    widths = {1: 22, 2: 10, 3: 22, 4: 30, 5: 22, 6: 18, 7: 22, 8: 40}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(issues_path)


# ---- Main ----------------------------------------------------------------


def derive_issues_path(annotated_path: Path) -> Path:
    """<name>_annotated.xlsx → <name>_issues.xlsx (sibling)."""
    stem = annotated_path.stem
    if not stem.endswith("_annotated"):
        raise SystemExit(
            "Expected an annotated file (filename should end in '_annotated.xlsx'). "
            f"Got: {annotated_path.name}"
        )
    base = stem[: -len("_annotated")]
    return annotated_path.with_name(base + "_issues.xlsx")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Apply Fill Value entries from the Missing Field Corrections tab "
            "of <name>_issues.xlsx to <name>_annotated.xlsx."
        )
    )
    parser.add_argument("annotated", help="Path to the <name>_annotated.xlsx file.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Compute changes but don't write any files.")
    args = parser.parse_args()

    annotated_path = Path(args.annotated).expanduser().resolve()
    if not annotated_path.exists():
        print(f"File not found: {annotated_path}", file=sys.stderr)
        return 1

    issues_path = derive_issues_path(annotated_path)
    if not issues_path.exists():
        print(
            f"Issues file not found: {issues_path}\n"
            "Run a scan first so the corrections worksheet is generated.",
            file=sys.stderr,
        )
        return 1

    rows = read_fill_rows(issues_path)
    if not rows:
        print("No missing-field rows found in the issues file — nothing to do.")
        return 0

    fills = [r for r in rows if r["fill_value"]]
    skipped = [r for r in rows if not r["fill_value"]]

    print(f"Missing-field rows total: {len(rows)}")
    print(f"  → fills to apply:        {len(fills)}")
    print(f"  → deferred (empty):      {len(skipped)}")

    if args.dry_run:
        print("\n[dry-run] would set these cells:")
        for f in fills:
            print(
                f"   row {f['excel_row']:>4}  col {f['column']:<25}  → {f['fill_value']!r}"
                f"   ({f['source'] or 'user-typed'})"
            )
        return 0

    audit, skipped_unknown_cols = apply_to_workbook(annotated_path, fills)
    print(f"\nCell fills written to {annotated_path.name}: {len(audit)}")
    if skipped_unknown_cols:
        print(f"  ! Skipped {len(skipped_unknown_cols)} fill(s) — column(s) not found in "
              f"tracks sheet: {sorted(set(skipped_unknown_cols))}")

    write_audit_log(issues_path, audit, skipped_unknown_cols, annotated_path)
    print(f"Audit log written to {issues_path.name} → 'Applied Missing Fields' tab.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
